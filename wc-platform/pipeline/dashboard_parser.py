"""
pipeline/dashboard_parser.py

Parses the real dashboard export - the file Mukesh downloads and then
hand-annotates with ACTION and SUGGESTED SCHEME columns before uploading.

The split that shapes this whole module: the DASHBOARD columns are
machine-generated and reliable; anything hand-added varies file to file.
So required columns are asserted loudly, hand-added ones are treated as
optional and their contents as untrusted vocabulary, and nothing that
looks wrong is ever silently dropped - it surfaces as a warning the RM
resolves on the review screen.

Everything is located BY HEADER NAME. Column order and presence differ
between files, so any positional read would work on the file it was
written against and quietly mis-parse the next one. The single exception
is the CAGR-shift repair, which deliberately reads one column to the
right of a named column - see _repair_cagr_shift.
"""

from __future__ import annotations

import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Optional, Union

from openpyxl import load_workbook

from pipeline.parser import Holding, Sip


# --------------------------------------------------------------------------
# Sheet detection
# --------------------------------------------------------------------------
# Matched as a case-insensitive SUBSTRING of the sheet name, because real
# files number their tabs inconsistently: "1. Mutual Fund" / "Mutual Fund",
# "Long Term" / "2. Long Term", "SIP" / "Sip" / "3. SIP".

SHEET_HOLDINGS = "mutual fund"
SHEET_ACTIONS = "long term"
SHEET_SIP = "sip"


# --------------------------------------------------------------------------
# Header normalisation and synonyms
# --------------------------------------------------------------------------

def normalise_header(value: Any) -> str:
    """Strip, uppercase, collapse internal whitespace. Applied to every
    header cell before anything looks at it, so "Folio  No " and
    "folio no" are the same column."""
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value).strip().upper())


# Different exports name the same column differently. Mapped AFTER
# normalisation so only one spelling of each has to be handled downstream.
HEADER_SYNONYMS = {
    "SUGGESTION": "SUGGESTED SCHEME",
    "ACTUAL CAPITAL GAIN": "GAIN",
    "PURCHASE AMOUNT": "PURCHASE VALUE",
    # The SIP sheets spell it "SCHEME NAME" while the holdings sheet says
    # "SCHEME". Without this every real file reported "the SIP sheet has
    # no SCHEME column" and silently read zero SIPs - MANISH's 95 SIPs,
    # including the SIP Stop instructions, never reached the report.
    "SCHEME NAME": "SCHEME",
}


def canonical_header(value: Any) -> str:
    header = normalise_header(value)
    return HEADER_SYNONYMS.get(header, header)


REQUIRED_HOLDING_COLUMNS = (
    "SCHEME",
    "FOLIO NO",
    "BALANCE UNITS",
    "PURCHASE VALUE",
    "CURRENT VALUE",
    "GAIN",
    "HOLDING DAYS",
    "ABSOLUTE RETURN (%)",
    "CAGR (%)",
)

OPTIONAL_HOLDING_COLUMNS = (
    "NAME",
    "CURRENT NAV",
    "RANK 6M", "RANK 1Y", "RANK 3Y", "RANK 5Y",
    "RETURN 6M", "RETURN 1Y", "RETURN 3Y", "RETURN 5Y",
)


# --------------------------------------------------------------------------
# Client vs asset-class header rows
# --------------------------------------------------------------------------
# Both render as "text in the first column, every other column empty".
# They are told apart by MEMBERSHIP OF THIS SET, never by looking for a
# PAN pattern: one real file carries "EXAMPLE HOLDINGS PVT LTD"
# whose PAN is truncated, and a PAN-shaped test would file that company
# under whichever client happened to precede it.

ASSET_CLASS_ROWS = {
    "EQUITY", "HYBRID", "DEBT", "OTHER", "GOLD", "LIQUID", "ARBITRAGE",
}

TOTAL_ROW_MARKERS = ("grand total", "mutual funds total", "sub total", "subtotal", "total")


# --------------------------------------------------------------------------
# Action vocabulary
# --------------------------------------------------------------------------
# Hand-typed, so matched case-insensitively on a normalised token. An
# action outside this vocabulary is NEVER dropped - it becomes a warning
# the RM sees and decides on, because a mis-spelled "swtich" silently
# discarded is a transaction that never happens.

TRANSACTION_ACTIONS = {"switch", "redeem", "trim"}
THINGS_TO_DO_ACTIONS = {"watchlist", "accumulate", "tax loss harvest"}


def classify_action(raw: Any) -> tuple[str, Optional[str]]:
    """Returns (kind, canonical) where kind is 'transaction',
    'things_to_do' or 'unrecognised'."""
    text = re.sub(r"\s+", " ", str(raw or "").strip().lower())
    if not text:
        return "none", None
    if text in TRANSACTION_ACTIONS:
        return "transaction", text
    if text in THINGS_TO_DO_ACTIONS:
        return "things_to_do", text
    return "unrecognised", text


# --------------------------------------------------------------------------
# The CAGR-shift defect
# --------------------------------------------------------------------------
# Some exported rows have the GAIN value duplicated into the CAGR column,
# which pushes the real CAGR one column to the right (seen in one real
# file, 6 rows: gain=185564.32, CAGR=185564.32, real CAGR=97.31).
#
# Two independent detectors, because either alone misses cases: CAGR
# essentially equal to the gain catches the duplication directly, and an
# implausible magnitude catches a shift whose gain happened to be small.

CAGR_GAIN_MATCH_TOLERANCE = 0.01     # within 1% of the gain value
MAX_PLAUSIBLE_CAGR_PCT = 150.0       # a fund CAGR beyond this is a data error


@dataclass
class ParseWarning:
    """Everything the parser could not resolve on its own. Carries enough
    location to point a human at the actual cell."""
    kind: str
    message: str
    sheet: Optional[str] = None
    row: Optional[int] = None
    column: Optional[str] = None
    client: Optional[str] = None

    def describe(self) -> str:
        where = []
        if self.sheet:
            where.append(f"sheet '{self.sheet}'")
        if self.row:
            where.append(f"row {self.row}")
        if self.column:
            where.append(f"column '{self.column}'")
        location = f" ({', '.join(where)})" if where else ""
        return f"{self.message}{location}"


class ParseError(Exception):
    """Blocks progress. Always names the sheet, and the row and column
    where those are known - an error the RM cannot locate is an error
    they cannot fix."""

    def __init__(self, message: str, sheet: Optional[str] = None,
                 row: Optional[int] = None, column: Optional[str] = None):
        self.sheet, self.row, self.column = sheet, row, column
        parts = [message]
        where = []
        if sheet:
            where.append(f"sheet '{sheet}'")
        if row:
            where.append(f"row {row}")
        if column:
            where.append(f"column '{column}'")
        if where:
            parts.append(f"({', '.join(where)})")
        super().__init__(" ".join(parts))


@dataclass
class ParsedAction:
    """One hand-added ACTION, matched back to the holding it applies to."""
    scheme: str
    folio: Optional[str]
    action_raw: str
    kind: str                      # transaction | things_to_do | unrecognised
    canonical: Optional[str]
    suggested_scheme: Optional[str] = None
    sheet: Optional[str] = None
    row: Optional[int] = None
    matched: bool = False


@dataclass
class ClientPortfolio:
    name: str
    pan: Optional[str] = None
    holdings: list = field(default_factory=list)
    sips: list = field(default_factory=list)
    actions: list = field(default_factory=list)
    warnings: list = field(default_factory=list)
    reported_grand_total: Optional[float] = None

    @property
    def computed_grand_total(self) -> float:
        return sum(h.current_value for h in self.holdings if h.current_value is not None)

    @property
    def total_purchase_value(self) -> float:
        return sum(h.purchase_value for h in self.holdings if h.purchase_value is not None)

    @property
    def grand_total_reconciles(self) -> Optional[bool]:
        """None when the sheet carried no Grand Total row to check against
        - unknown is not the same as passing."""
        if self.reported_grand_total is None:
            return None
        return abs(self.reported_grand_total - self.computed_grand_total) <= RECONCILE_TOLERANCE_RUPEES


RECONCILE_TOLERANCE_RUPEES = 1.0


@dataclass
class WorkbookParseResult:
    clients: list = field(default_factory=list)
    warnings: list = field(default_factory=list)
    source_name: str = ""
    sheet_names: list = field(default_factory=list)
    # The single workbook-level "Grand Total" figure, when the sheet had
    # one. Distinct from any client's own total row.
    reported_workbook_total: Optional[float] = None

    @property
    def computed_workbook_total(self) -> float:
        return sum(c.computed_grand_total for c in self.clients)

    @property
    def workbook_total_reconciles(self) -> Optional[bool]:
        if self.reported_workbook_total is None:
            return None
        return abs(self.reported_workbook_total - self.computed_workbook_total) <= RECONCILE_TOLERANCE_RUPEES

    def client_names(self) -> list:
        return [c.name for c in self.clients]

    def client(self, name: str) -> Optional[ClientPortfolio]:
        return next((c for c in self.clients if c.name == name), None)


# --------------------------------------------------------------------------
# Cell helpers
# --------------------------------------------------------------------------

def _to_float(value: Any) -> Optional[float]:
    """Never guesses. A cell that isn't a number becomes None and the
    caller decides whether that is a warning or normal."""
    if value is None:
        return None
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return float(value)
    text = str(value).strip().replace(",", "").replace("₹", "").replace("%", "")
    if text in ("", "-", "--", "N/A", "NA", "n/a"):
        return None
    try:
        return float(text)
    except ValueError:
        return None


def _text(value: Any) -> str:
    return "" if value is None else str(value).strip()


def _is_blank_row(values: list) -> bool:
    return all(_text(v) == "" for v in values)


def _is_total_row(first_cell: Any) -> bool:
    text = _text(first_cell).lower()
    return any(marker in text for marker in TOTAL_ROW_MARKERS)


def _looks_like_group_header(values: list) -> bool:
    """Text in the first column and nothing anywhere else."""
    if not values or _text(values[0]) == "":
        return False
    return all(_text(v) == "" for v in values[1:])


CLIENT_PAN_RE = re.compile(r"^(?P<name>.+?)\s*\((?P<pan>[^)]*)\)\s*$")


def split_client_label(label: str) -> tuple[str, Optional[str]]:
    """'PRIYA SHARMA (AAAAA1111A)' -> ('PRIYA SHARMA', 'AAAAA1111A').

    Whatever is inside the brackets is taken as the PAN verbatim, valid
    or not. One real file carries a truncated PAN, and rejecting it for
    failing a PAN regex would drop the client entirely rather than
    recording an imperfect identifier.
    """
    match = CLIENT_PAN_RE.match(label.strip())
    if not match:
        return label.strip(), None
    pan = match.group("pan").strip() or None
    return match.group("name").strip(), pan


# --------------------------------------------------------------------------
# Sheet location
# --------------------------------------------------------------------------

def find_sheet(workbook, needle: str):
    for name in workbook.sheetnames:
        if needle in name.strip().lower():
            return workbook[name]
    return None


def _header_index(values: list) -> dict:
    """{canonical header -> column index}. Later duplicates do not
    overwrite earlier ones, so a stray repeated header can't silently
    redirect a required column."""
    index = {}
    for position, cell in enumerate(values):
        header = canonical_header(cell)
        if header and header not in index:
            index[header] = position
    return index


def _find_header_row(worksheet, required: tuple, sheet_name: str, max_scan: int = 30) -> tuple:
    """Locates the header row by looking for the required columns rather
    than assuming row 1 - real exports carry title and filter rows above
    the table, and the count varies."""
    for row_idx, row in enumerate(
        worksheet.iter_rows(min_row=1, max_row=max_scan, values_only=True), start=1
    ):
        values = list(row)
        index = _header_index(values)
        if all(column in index for column in required):
            return row_idx, index
    # Nothing matched - report what WAS found on the best candidate row so
    # the message names actual columns instead of just "not found".
    best_row, best_index, best_score = None, {}, -1
    for row_idx, row in enumerate(
        worksheet.iter_rows(min_row=1, max_row=max_scan, values_only=True), start=1
    ):
        index = _header_index(list(row))
        score = sum(1 for column in required if column in index)
        if score > best_score:
            best_row, best_index, best_score = row_idx, index, score
    missing = [column for column in required if column not in best_index]
    raise ParseError(
        f"could not find the header row. Required column(s) {missing} are missing. "
        f"Closest match had: {sorted(best_index)}",
        sheet=sheet_name, row=best_row,
    )


# --------------------------------------------------------------------------
# CAGR-shift repair
# --------------------------------------------------------------------------

def _cagr_looks_shifted(gain: Optional[float], cagr: Optional[float]) -> Optional[str]:
    """Returns the reason the CAGR is suspect, or None if it looks fine."""
    if cagr is None:
        return None
    if gain not in (None, 0) and abs(cagr - gain) <= abs(gain) * CAGR_GAIN_MATCH_TOLERANCE:
        return "CAGR equals the gain value"
    if abs(cagr) > MAX_PLAUSIBLE_CAGR_PCT:
        return f"CAGR of {cagr:,.2f}% is not plausible"
    return None


def _repair_cagr_shift(row_values: list, cagr_position: int, gain: Optional[float],
                       cagr: Optional[float], scheme: str, sheet: str, row_idx: int,
                       warnings: list) -> Optional[float]:
    """Handles the known export defect where GAIN is duplicated into the
    CAGR column, pushing the real CAGR one column right.

    This is the ONE place that reads by position rather than by header,
    and deliberately so: the defect IS a positional shift, so the repair
    has to look at the neighbouring cell. It only ever fires on a row
    already flagged as suspect.

    If the neighbour holds a plausible CAGR the row is corrected. If it
    does not, the CAGR is discarded (None) rather than ingested, and a
    warning names the row. The holding itself is KEPT: dropping it would
    silently break the grand-total reconciliation this same parser
    asserts, turning a bad cell into a missing lakh.
    """
    reason = _cagr_looks_shifted(gain, cagr)
    if reason is None:
        return cagr

    neighbour = _to_float(row_values[cagr_position + 1]) if cagr_position + 1 < len(row_values) else None
    if neighbour is not None and _cagr_looks_shifted(gain, neighbour) is None:
        warnings.append(ParseWarning(
            kind="cagr_shift_corrected",
            message=(f"'{scheme}': {reason} ({cagr:,.2f}); the real CAGR was one column right "
                     f"and has been read as {neighbour:,.2f}%"),
            sheet=sheet, row=row_idx, column="CAGR (%)",
        ))
        return neighbour

    warnings.append(ParseWarning(
        kind="cagr_rejected",
        message=(f"'{scheme}': {reason} ({cagr:,.2f}) and no usable value in the next column. "
                 f"CAGR discarded for this holding; every other figure on the row is unaffected."),
        sheet=sheet, row=row_idx, column="CAGR (%)",
    ))
    return None


# --------------------------------------------------------------------------
# Holdings sheet
# --------------------------------------------------------------------------

def _parse_holdings_sheet(worksheet, warnings: list, workbook_totals: Optional[dict] = None) -> list:
    """Returns [ClientPortfolio] with holdings attached.

    Grouping is by NAME column when the sheet has one, and by the
    client header rows otherwise - the two layouts both occur.
    """
    sheet = worksheet.title
    header_row, index = _find_header_row(worksheet, REQUIRED_HOLDING_COLUMNS, sheet)

    missing = [c for c in REQUIRED_HOLDING_COLUMNS if c not in index]
    if missing:  # pragma: no cover - _find_header_row already guarantees this
        raise ParseError(f"required column(s) missing: {missing}", sheet=sheet, row=header_row)

    has_name_column = "NAME" in index
    clients: list = []
    by_name: dict = {}
    current: Optional[ClientPortfolio] = None
    current_category: Optional[str] = None

    def client_for(label: str, pan: Optional[str] = None) -> ClientPortfolio:
        if label not in by_name:
            portfolio = ClientPortfolio(name=label, pan=pan)
            by_name[label] = portfolio
            clients.append(portfolio)
        elif pan and not by_name[label].pan:
            by_name[label].pan = pan
        return by_name[label]

    max_column = max(index.values()) + 2   # +2 so the shift repair can look right
    for row_idx, row in enumerate(
        worksheet.iter_rows(min_row=header_row + 1, values_only=True), start=header_row + 1
    ):
        values = list(row)
        if _is_blank_row(values):
            continue

        first = _text(values[0])

        # Total rows are captured for the reconciliation check, then
        # skipped as data.
        #
        # These exports carry TWO kinds, and conflating them produced a
        # false mismatch on every multi-client file: a per-client row
        # ("ARJUN MEHTA Total") that totals just that client, and
        # ONE workbook-level "Grand Total" that totals every client in the
        # sheet. Attaching the workbook figure to whichever client
        # happened to be open last compared one client's holdings against
        # the whole book - on MANISH's file that read as a Rs 28.16 crore
        # discrepancy when nothing was actually wrong.
        if _is_total_row(first):
            total = (_to_float(values[index["CURRENT VALUE"]])
                     if index["CURRENT VALUE"] < len(values) else None)
            if total is not None:
                lowered = first.lower()
                if lowered.startswith("grand total"):
                    workbook_totals["grand_total"] = total
                elif current is not None and lowered.endswith("total"):
                    # "<CLIENT NAME> Total" - belongs to the open client.
                    current.reported_grand_total = total
            continue

        if _looks_like_group_header(values):
            label = first
            if normalise_header(label) in ASSET_CLASS_ROWS:
                current_category = label.strip().title()
            else:
                name, pan = split_client_label(label)
                current = client_for(name, pan)
                current_category = None
            continue

        scheme = _text(values[index["SCHEME"]]) if index["SCHEME"] < len(values) else ""
        if not scheme:
            continue

        if has_name_column:
            raw_name = _text(values[index["NAME"]]) if index["NAME"] < len(values) else ""
            if raw_name:
                name, pan = split_client_label(raw_name)
                current = client_for(name, pan)

        if current is None:
            # A data row before any client header and with no NAME column
            # means the grouping cannot be established - refusing here is
            # better than filing someone's holdings under a guess.
            raise ParseError(
                f"holding '{scheme}' appears before any client header row, and the sheet has no "
                f"NAME column to group by, so it cannot be attributed to a client",
                sheet=sheet, row=row_idx, column="SCHEME",
            )

        gain = _to_float(values[index["GAIN"]]) if index["GAIN"] < len(values) else None
        cagr_position = index["CAGR (%)"]
        cagr = _to_float(values[cagr_position]) if cagr_position < len(values) else None
        cagr = _repair_cagr_shift(values, cagr_position, gain, cagr, scheme, sheet, row_idx,
                                  current.warnings)

        def column(name: str):
            position = index.get(name)
            if position is None or position >= len(values):
                return None
            return values[position]

        current.holdings.append(Holding(
            member=current.name,
            pan=current.pan,
            category=current_category,
            scheme=scheme,
            folio=_text(column("FOLIO NO")) or None,
            balance_units=_to_float(column("BALANCE UNITS")),
            purchase_value=_to_float(column("PURCHASE VALUE")),
            current_value=_to_float(column("CURRENT VALUE")),
            gain=gain,
            absolute_return_pct=_to_float(column("ABSOLUTE RETURN (%)")),
            cagr_pct=cagr,
            source_sheet=sheet,
            source_row=row_idx,
        ))

    return clients


# --------------------------------------------------------------------------
# Actions sheet
# --------------------------------------------------------------------------

def _match_key(scheme: Any, folio: Any) -> tuple:
    return (
        re.sub(r"\s+", " ", _text(scheme)).lower(),
        _text(folio).lower() or None,
    )


def _parse_actions_sheet(worksheet, clients: list, warnings: list) -> None:
    """Reads hand-added ACTION / SUGGESTED SCHEME and attaches each to the
    holding it matches on (scheme, folio).

    A sheet with no ACTION column, or with every ACTION blank, is normal -
    one real client has no actions at all, and that must still produce a
    valid report with an empty Transaction Snapshot.
    """
    sheet = worksheet.title
    try:
        header_row, index = _find_header_row(worksheet, ("SCHEME",), sheet)
    except ParseError:
        warnings.append(ParseWarning(
            kind="actions_sheet_unreadable",
            message="the long-term sheet has no SCHEME column, so no actions were read from it",
            sheet=sheet,
        ))
        return

    if "ACTION" not in index:
        warnings.append(ParseWarning(
            kind="no_action_column",
            message="the long-term sheet has no ACTION column - no transactions will be proposed",
            sheet=sheet,
        ))
        return

    lookup: dict = {}
    for client in clients:
        for holding in client.holdings:
            lookup.setdefault(_match_key(holding.scheme, holding.folio), (client, holding))
            # Folio-less fallback, so a blank folio on the action sheet
            # still finds a uniquely-named scheme.
            lookup.setdefault((_match_key(holding.scheme, None)[0], None), (client, holding))

    for row_idx, row in enumerate(
        worksheet.iter_rows(min_row=header_row + 1, values_only=True), start=header_row + 1
    ):
        values = list(row)
        if _is_blank_row(values) or _is_total_row(_text(values[0])):
            continue

        def column(name: str):
            position = index.get(name)
            if position is None or position >= len(values):
                return None
            return values[position]

        raw_action = _text(column("ACTION"))
        if not raw_action:
            continue

        scheme = _text(column("SCHEME"))
        folio = _text(column("FOLIO NO")) or None
        kind, canonical = classify_action(raw_action)

        action = ParsedAction(
            scheme=scheme, folio=folio, action_raw=raw_action, kind=kind, canonical=canonical,
            suggested_scheme=_text(column("SUGGESTED SCHEME")) or None,
            sheet=sheet, row=row_idx,
        )

        target = lookup.get(_match_key(scheme, folio)) or lookup.get((_match_key(scheme, None)[0], None))
        if target is None:
            warnings.append(ParseWarning(
                kind="action_unmatched",
                message=(f"action '{raw_action}' on '{scheme}' (folio {folio or 'blank'}) does not "
                         f"match any holding, so it has been left out of the report"),
                sheet=sheet, row=row_idx, column="ACTION",
            ))
            continue

        client, holding = target
        action.matched = True
        client.actions.append(action)

        if kind == "unrecognised":
            client.warnings.append(ParseWarning(
                kind="action_unrecognised",
                message=(f"'{scheme}': action '{raw_action}' is not a recognised instruction. "
                         f"Known: {', '.join(sorted(TRANSACTION_ACTIONS | THINGS_TO_DO_ACTIONS))}. "
                         f"It has NOT been applied - confirm what was intended."),
                sheet=sheet, row=row_idx, column="ACTION", client=client.name,
            ))
        else:
            holding.action = canonical
            holding.suggested_scheme = action.suggested_scheme


# --------------------------------------------------------------------------
# SIP sheet
# --------------------------------------------------------------------------

def _parse_sip_sheet(worksheet, clients: list, warnings: list) -> None:
    sheet = worksheet.title
    try:
        header_row, index = _find_header_row(worksheet, ("SCHEME",), sheet)
    except ParseError:
        warnings.append(ParseWarning(
            kind="sip_sheet_unreadable",
            message="the SIP sheet has no SCHEME column, so no SIPs were read",
            sheet=sheet,
        ))
        return

    amount_column = next(
        (c for c in ("SIP AMOUNT", "AMOUNT", "INSTALMENT AMOUNT", "INSTALLMENT AMOUNT")
         if c in index), None
    )
    by_name = {c.name: c for c in clients}
    current = clients[0] if len(clients) == 1 else None

    for row_idx, row in enumerate(
        worksheet.iter_rows(min_row=header_row + 1, values_only=True), start=header_row + 1
    ):
        values = list(row)
        if _is_blank_row(values) or _is_total_row(_text(values[0])):
            continue

        if _looks_like_group_header(values):
            name, _pan = split_client_label(_text(values[0]))
            if normalise_header(name) not in ASSET_CLASS_ROWS:
                current = by_name.get(name, current)
            continue

        def column(name: Optional[str]):
            if name is None:
                return None
            position = index.get(name)
            if position is None or position >= len(values):
                return None
            return values[position]

        scheme = _text(column("SCHEME"))
        if not scheme:
            continue

        # The SIP sheets attribute per ROW via their own client column,
        # not via group-header rows: every real file spells it "CLIENT"
        # (with a leading space in the cell, which normalisation strips).
        # Without reading it, every SIP row fell through to "could not be
        # attributed" and was skipped - 95 of them on MANISH's file,
        # including the SIP Stop instructions that drive the Mind Map.
        for client_column in ("CLIENT", "NAME"):
            if client_column in index:
                raw_name = _text(column(client_column))
                if raw_name:
                    name, _pan = split_client_label(raw_name)
                    matched = by_name.get(name)
                    if matched is None:
                        warnings.append(ParseWarning(
                            kind="sip_unknown_client",
                            message=(f"SIP row names client '{name}', which has no holdings in "
                                     f"this workbook - the SIP was skipped"),
                            sheet=sheet, row=row_idx,
                        ))
                        current = None
                    else:
                        current = matched
                break

        if current is None:
            warnings.append(ParseWarning(
                kind="sip_unattributed",
                message=f"SIP for '{scheme}' could not be attributed to a client and was skipped",
                sheet=sheet, row=row_idx,
            ))
            continue

        current.sips.append(Sip(
            member=current.name, pan=current.pan, scheme=scheme,
            sip_amount=_to_float(column(amount_column)),
            sip_date=_text(column("SIP DATE")) or _text(column("DATE")) or None,
            # The SIP sheet carries its own hand-added ACTION column
            # ("Stop"), which is what turns a running SIP into a SIP Stop
            # on the Mind Map.
            instruction=(_text(column("INSTRUCTION")) or _text(column("ACTION")) or None),
            source_sheet=sheet, source_row=row_idx,
        ))


# --------------------------------------------------------------------------
# Entry point
# --------------------------------------------------------------------------

def parse_dashboard_workbook(path: Union[str, Path]) -> WorkbookParseResult:
    """Parses one uploaded dashboard export into per-client portfolios.

    read_only + data_only are both mandatory: data_only=False returns
    formula strings instead of values, and read_only keeps a 300-row,
    multi-sheet workbook from being fully materialised.
    """
    path = Path(path)
    workbook = load_workbook(filename=str(path), read_only=True, data_only=True)
    try:
        warnings: list = []
        holdings_sheet = find_sheet(workbook, SHEET_HOLDINGS)
        if holdings_sheet is None:
            raise ParseError(
                f"no holdings sheet found. Expected a sheet whose name contains "
                f"'{SHEET_HOLDINGS}' (case-insensitive); this file has: {workbook.sheetnames}"
            )

        workbook_totals: dict = {}
        clients = _parse_holdings_sheet(holdings_sheet, warnings, workbook_totals)
        if not clients:
            raise ParseError(
                "no clients found - the holdings sheet has no client header rows and no NAME column",
                sheet=holdings_sheet.title,
            )

        actions_sheet = find_sheet(workbook, SHEET_ACTIONS)
        if actions_sheet is not None:
            _parse_actions_sheet(actions_sheet, clients, warnings)
        else:
            warnings.append(ParseWarning(
                kind="no_actions_sheet",
                message=(f"no sheet name contains '{SHEET_ACTIONS}', so no actions were read. "
                         f"The report will build with an empty Transaction Snapshot."),
            ))

        sip_sheet = find_sheet(workbook, SHEET_SIP)
        if sip_sheet is not None:
            _parse_sip_sheet(sip_sheet, clients, warnings)

        # Reconciliation runs at BOTH levels the sheet actually provides:
        # each client against its own "<NAME> Total" row, and the whole
        # book against the single "Grand Total" row. The workbook check is
        # the one that catches a client being dropped entirely - every
        # per-client total can reconcile while a missing client leaves the
        # book short.
        book_total = workbook_totals.get("grand_total")
        if book_total is not None:
            computed_book = sum(c.computed_grand_total for c in clients)
            if abs(computed_book - book_total) > RECONCILE_TOLERANCE_RUPEES:
                warnings.append(ParseWarning(
                    kind="workbook_total_mismatch",
                    message=(f"holdings across all {len(clients)} clients sum to "
                             f"Rs {computed_book:,.2f} but the sheet's Grand Total says "
                             f"Rs {book_total:,.2f} - a difference of "
                             f"Rs {computed_book - book_total:,.2f}"),
                    sheet=holdings_sheet.title,
                ))

        for client in clients:
            reconciles = client.grand_total_reconciles
            if reconciles is False:
                client.warnings.append(ParseWarning(
                    kind="grand_total_mismatch",
                    message=(f"holdings sum to Rs {client.computed_grand_total:,.2f} but the sheet's "
                             f"'{client.name} Total' row says Rs {client.reported_grand_total:,.2f} - a difference "
                             f"of Rs {client.computed_grand_total - client.reported_grand_total:,.2f}"),
                    sheet=holdings_sheet.title, client=client.name,
                ))

        return WorkbookParseResult(
            clients=clients, warnings=warnings,
            source_name=path.name, sheet_names=list(workbook.sheetnames),
            reported_workbook_total=book_total,
        )
    finally:
        workbook.close()
