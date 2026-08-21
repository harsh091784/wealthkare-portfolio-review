"""
pipeline/parser.py

Excel client-portfolio parser for the WC Securities / Wealthkare
Portfolio Review pipeline.

Parses a client portfolio workbook (as exported by the internal ops team)
into a structured, member-aware result: holdings, recommended transactions,
SIP details, and a warnings log for anything that could not be read safely.

Design notes
------------
- Workbooks are opened with data_only=True so formulas resolve to their last
  computed value instead of the formula string. This is mandatory — without
  it, numeric cells come back as "=SUM(...)"-style strings.
- Layouts vary per client (rows shift, extra blank rows appear, member
  blocks are different lengths), so this file NEVER assumes a fixed row
  offset for data. Instead it walks each sheet row-by-row as a small state
  machine, tracking "current member" and "current category" as it goes.
- Nothing is ever inferred or guessed. A missing/blank/malformed value is
  recorded as a warning and left out (or set to None) rather than filled in.
"""

from __future__ import annotations

import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Optional, Union

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

# --------------------------------------------------------------------------
# Constants
# --------------------------------------------------------------------------

# Sheet-name candidates, checked in order (case-insensitive substring match).
HOLDINGS_SHEET_CANDIDATES = ["mutual fund", "mutual funds", "holdings", "portfolio"]
RECOMMENDATION_SHEET_CANDIDATES = ["long term", "transaction snapshot", "recommendation", "recommendations"]
SIP_SHEET_CANDIDATES = ["sip", "sheet1"]

# Category headers that can appear embedded as a data row in the sheet.
CATEGORY_KEYWORDS = {
    "equity": "Equity",
    "hybrid": "Hybrid",
    "debt": "Debt",
    "other": "Other",
    "solution oriented": "Other",
    "fof": "Other",
}

# Scheme-name values that mark a totals row — must be filtered out before
# any total is computed, and are never treated as a real holding.
TOTAL_ROW_NAMES = {
    "grand total",
    "mutual funds total",
    "sub total",
    "subtotal",
    "total",
}

# A row is treated as a "member header" row when it matches this PAN pattern
# (10 chars: 5 letters, 4 digits, 1 letter) anywhere in the row, or when a
# cell literally starts with "Name:" / "PAN:".
PAN_REGEX = re.compile(r"\b[A-Z]{5}[0-9]{4}[A-Z]\b")


# --------------------------------------------------------------------------
# Data model
# --------------------------------------------------------------------------

@dataclass
class Holding:
    member: str
    pan: Optional[str]
    category: Optional[str]
    scheme: Optional[str]
    folio: Optional[str]
    balance_units: Optional[float]
    purchase_value: Optional[float]
    current_value: Optional[float]
    gain: Optional[float]
    absolute_return_pct: Optional[float]
    cagr_pct: Optional[float]
    action: Optional[str] = None            # e.g. "Switch", "Redeem", "Fresh Purchase"
    suggested_scheme: Optional[str] = None  # for recommendation rows
    source_sheet: Optional[str] = None
    source_row: Optional[int] = None


@dataclass
class Sip:
    member: str
    pan: Optional[str]
    scheme: Optional[str]
    sip_amount: Optional[float]
    sip_date: Optional[str]
    instruction: Optional[str] = None  # e.g. "Stop", "Start", "Continue"
    source_sheet: Optional[str] = None
    source_row: Optional[int] = None


@dataclass
class MemberResult:
    name: str
    pan: Optional[str] = None
    holdings: list[Holding] = field(default_factory=list)
    sips: list[Sip] = field(default_factory=list)
    total_purchase_value: float = 0.0
    total_current_value: float = 0.0
    total_gain: float = 0.0


@dataclass
class PortfolioParseResult:
    members: list[MemberResult] = field(default_factory=list)
    holdings: list[Holding] = field(default_factory=list)          # flat, all members
    sips: list[Sip] = field(default_factory=list)                  # flat, all members
    consolidated_purchase_value: float = 0.0
    consolidated_current_value: float = 0.0
    consolidated_gain: float = 0.0
    warnings: list[str] = field(default_factory=list)

    def member_by_name(self, name: str) -> Optional[MemberResult]:
        for m in self.members:
            if m.name == name:
                return m
        return None


# --------------------------------------------------------------------------
# Helpers
# --------------------------------------------------------------------------

def _norm(value: Any) -> str:
    """Stripped, lowercased string form of a cell value ('' for None)."""
    if value is None:
        return ""
    return str(value).strip().lower()


def _is_total_row(scheme_name: Any) -> bool:
    return _norm(scheme_name) in TOTAL_ROW_NAMES


def _to_float(value: Any) -> Optional[float]:
    """Best-effort numeric coercion. Returns None (never a guess) on failure."""
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace(",", "")
    if text in ("", "-", "NA", "N/A"):
        return None
    try:
        return float(text)
    except ValueError:
        return None


def _row_values(ws: Worksheet, row_idx: int, max_col: int) -> list[Any]:
    return [ws.cell(row=row_idx, column=c).value for c in range(1, max_col + 1)]


def _row_text_blob(values: list[Any]) -> str:
    """All non-empty cells in a row joined into one lowercase string, for
    keyword scanning (category headers, member headers, etc.)."""
    return " ".join(_norm(v) for v in values if v is not None and _norm(v) != "")


def _find_sheet(wb, candidates: list[str]) -> Optional[Worksheet]:
    """Case-insensitive substring match against sheet names, in candidate order."""
    sheet_names = wb.sheetnames
    for cand in candidates:
        for name in sheet_names:
            if cand in name.lower():
                return wb[name]
    return None


def _detect_category(row_blob: str) -> Optional[str]:
    for keyword, label in CATEGORY_KEYWORDS.items():
        if keyword in row_blob:
            return label
    return None


def _extract_pan(row_blob_raw: list[Any]) -> Optional[str]:
    for v in row_blob_raw:
        if v is None:
            continue
        m = PAN_REGEX.search(str(v).upper())
        if m:
            return m.group(0)
    return None


def _looks_like_member_header(values: list[Any], row_blob: str) -> bool:
    """A row is a member header if it carries a PAN, or an explicit
    'Name:' / 'Client Name' style label."""
    if PAN_REGEX.search(row_blob.upper()):
        return True
    if "name:" in row_blob or "client name" in row_blob or "pan:" in row_blob:
        return True
    return False


def _extract_member_name(values: list[Any]) -> Optional[str]:
    """Pull a plausible member name out of a header row. Prefers the text
    following a 'Name:' label; falls back to the first non-empty string
    cell that isn't itself a PAN."""
    for v in values:
        if v is None:
            continue
        text = str(v).strip()
        low = text.lower()
        if low.startswith("name:"):
            candidate = text.split(":", 1)[1].strip()
            if candidate:
                return candidate
        if low.startswith("client name:"):
            candidate = text.split(":", 1)[1].strip()
            if candidate:
                return candidate
    for v in values:
        if v is None:
            continue
        text = str(v).strip()
        if not text:
            continue
        if PAN_REGEX.fullmatch(text.upper()):
            continue
        if ":" in text:
            continue
        return text
    return None


# --------------------------------------------------------------------------
# Holdings sheet parser ("Mutual Fund" / holdings-style layout)
# --------------------------------------------------------------------------

# Column header keywords we look for once, in the sheet's own header row,
# so we're tolerant of column order differing between exports. We still
# never assume a *row* offset — only which *column* a value lives in, and
# even that is discovered dynamically per sheet.
HOLDINGS_HEADER_MAP = {
    "scheme": ["scheme name", "scheme", "fund name", "fund"],
    "folio": ["folio no", "folio"],
    "balance_units": ["balance units", "units", "unit balance"],
    "purchase_value": ["purchase value", "invested value", "investment value", "cost value"],
    "current_value": ["current value", "market value", "valuation"],
    "gain": ["gain", "gain/loss", "unrealised gain", "unrealized gain"],
    "absolute_return_pct": ["absolute return", "abs return", "return %"],
    "cagr_pct": ["cagr", "xirr"],
    "action": ["action", "recommendation", "suggested action"],
    "suggested_scheme": ["suggested scheme", "switch to", "recommended scheme"],
}


def _build_header_index(values: list[Any]) -> dict[str, int]:
    """Given a candidate header row, map our canonical field names to the
    0-based column index whose header text matches."""
    index_map: dict[str, int] = {}
    for col_idx, cell in enumerate(values):
        text = _norm(cell)
        if not text:
            continue
        for field_name, keywords in HOLDINGS_HEADER_MAP.items():
            if field_name in index_map:
                continue
            for kw in keywords:
                if kw in text:
                    index_map[field_name] = col_idx
                    break
    return index_map


def _looks_like_header_row(values: list[Any]) -> bool:
    blob = _row_text_blob(values)
    hits = 0
    for keywords in HOLDINGS_HEADER_MAP.values():
        if any(kw in blob for kw in keywords):
            hits += 1
    return hits >= 2  # a real header row will match several canonical fields at once


def _parse_holdings_or_recommendations_sheet(
    ws: Worksheet,
    warnings: list[str],
    default_sheet_label: str,
) -> list[Holding]:
    holdings: list[Holding] = []
    if ws is None:
        return holdings

    max_col = ws.max_column or 1
    header_index: dict[str, int] = {}
    current_member = "Unknown Member"
    current_pan: Optional[str] = None
    current_category: Optional[str] = None

    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_col=max_col, values_only=True), start=1):
        values = list(row)
        if all(v is None or _norm(v) == "" for v in values):
            continue  # skip fully blank rows

        blob = _row_text_blob(values)

        # 1) Member header row?
        if _looks_like_member_header(values, blob):
            name = _extract_member_name(values)
            pan = _extract_pan(values)
            if name:
                current_member = name
            if pan:
                current_pan = pan
            elif not name:
                warnings.append(
                    f"[{ws.title}] row {row_idx}: looked like a member header but no name/PAN could be parsed."
                )
            continue

        # 2) Category header embedded as a data row?
        detected_category = _detect_category(blob)
        if detected_category and not _looks_like_header_row(values):
            # Guard: a header row can legitimately contain the word "equity"
            # inside "Equity Scheme Name" etc. Only treat as a pure category
            # marker if the row doesn't also look like a column header AND
            # doesn't have a numeric current-value-like cell (i.e. it's a
            # section banner, not a data row).
            numeric_cells = [v for v in values if isinstance(v, (int, float))]
            if not numeric_cells:
                current_category = detected_category
                continue

        # 3) Column header row -> (re)build header index and move on.
        if _looks_like_header_row(values):
            header_index = _build_header_index(values)
            continue

        # 4) Total row -> skip entirely, never counted as a holding.
        scheme_col = header_index.get("scheme", 0)
        scheme_val = values[scheme_col] if scheme_col < len(values) else None
        if _is_total_row(scheme_val):
            continue

        # 5) Otherwise: treat as a data row IF we have at least discovered
        # a scheme column and this row has a non-empty scheme name.
        if not header_index:
            # No header seen yet on this sheet - nothing to safely map to.
            continue
        if scheme_val is None or _norm(scheme_val) == "":
            continue

        def get(field_name: str) -> Any:
            idx = header_index.get(field_name)
            if idx is None or idx >= len(values):
                return None
            return values[idx]

        folio = get("folio")
        balance_units = _to_float(get("balance_units"))
        purchase_value = _to_float(get("purchase_value"))
        current_value = _to_float(get("current_value"))
        gain = _to_float(get("gain"))
        abs_return = _to_float(get("absolute_return_pct"))
        cagr = _to_float(get("cagr_pct"))
        action = get("action")
        suggested_scheme = get("suggested_scheme")

        # Never infer: log a warning for anything expected-but-missing on a
        # row that otherwise looks like a genuine holding line.
        if purchase_value is None:
            warnings.append(
                f"[{ws.title}] row {row_idx} ({scheme_val}): purchase value missing/blank/malformed."
            )
        if current_value is None:
            warnings.append(
                f"[{ws.title}] row {row_idx} ({scheme_val}): current value missing/blank/malformed."
            )

        holdings.append(
            Holding(
                member=current_member,
                pan=current_pan,
                category=current_category,
                scheme=str(scheme_val).strip(),
                folio=str(folio).strip() if folio is not None else None,
                balance_units=balance_units,
                purchase_value=purchase_value,
                current_value=current_value,
                gain=gain,
                absolute_return_pct=abs_return,
                cagr_pct=cagr,
                action=str(action).strip() if action not in (None, "") else None,
                suggested_scheme=str(suggested_scheme).strip() if suggested_scheme not in (None, "") else None,
                source_sheet=ws.title if ws.title else default_sheet_label,
                source_row=row_idx,
            )
        )

    return holdings


# --------------------------------------------------------------------------
# SIP sheet parser
# --------------------------------------------------------------------------

SIP_HEADER_MAP = {
    "scheme": ["scheme name", "scheme", "fund name", "fund"],
    "sip_amount": ["sip amount", "amount", "installment amount"],
    "sip_date": ["sip date", "date", "installment date"],
    "instruction": ["instruction", "action", "status", "stop/start", "remarks"],
}


def _build_sip_header_index(values: list[Any]) -> dict[str, int]:
    index_map: dict[str, int] = {}
    for col_idx, cell in enumerate(values):
        text = _norm(cell)
        if not text:
            continue
        for field_name, keywords in SIP_HEADER_MAP.items():
            if field_name in index_map:
                continue
            for kw in keywords:
                if kw in text:
                    index_map[field_name] = col_idx
                    break
    return index_map


def _looks_like_sip_header_row(values: list[Any]) -> bool:
    blob = _row_text_blob(values)
    hits = 0
    for keywords in SIP_HEADER_MAP.values():
        if any(kw in blob for kw in keywords):
            hits += 1
    return hits >= 2


def _parse_sip_sheet(ws: Worksheet, warnings: list[str]) -> list[Sip]:
    sips: list[Sip] = []
    if ws is None:
        return sips

    max_col = ws.max_column or 1
    header_index: dict[str, int] = {}
    current_member = "Unknown Member"
    current_pan: Optional[str] = None

    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_col=max_col, values_only=True), start=1):
        values = list(row)
        if all(v is None or _norm(v) == "" for v in values):
            continue

        blob = _row_text_blob(values)

        if _looks_like_member_header(values, blob):
            name = _extract_member_name(values)
            pan = _extract_pan(values)
            if name:
                current_member = name
            if pan:
                current_pan = pan
            continue

        if _looks_like_sip_header_row(values):
            header_index = _build_sip_header_index(values)
            continue

        scheme_col = header_index.get("scheme", 0)
        scheme_val = values[scheme_col] if scheme_col < len(values) else None
        if _is_total_row(scheme_val):
            continue
        if not header_index or scheme_val is None or _norm(scheme_val) == "":
            continue

        def get(field_name: str) -> Any:
            idx = header_index.get(field_name)
            if idx is None or idx >= len(values):
                return None
            return values[idx]

        sip_amount = _to_float(get("sip_amount"))
        sip_date = get("sip_date")
        instruction = get("instruction")

        if sip_amount is None:
            warnings.append(
                f"[{ws.title}] row {row_idx} ({scheme_val}): SIP amount missing/blank/malformed."
            )

        sips.append(
            Sip(
                member=current_member,
                pan=current_pan,
                scheme=str(scheme_val).strip(),
                sip_amount=sip_amount,
                sip_date=str(sip_date).strip() if sip_date is not None else None,
                instruction=str(instruction).strip() if instruction not in (None, "") else None,
                source_sheet=ws.title,
                source_row=row_idx,
            )
        )

    return sips


# --------------------------------------------------------------------------
# Top-level entry point
# --------------------------------------------------------------------------

def parse_portfolio_workbook(path: Union[str, Path]) -> PortfolioParseResult:
    """Parse a client portfolio workbook into a PortfolioParseResult.

    Never raises on malformed/missing data cells — those become warnings.
    Will raise if the file itself cannot be opened (bad path / not xlsx).
    """
    path = Path(path)
    warnings: list[str] = []

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)

    holdings_ws = _find_sheet(wb, HOLDINGS_SHEET_CANDIDATES)
    recs_ws = _find_sheet(wb, RECOMMENDATION_SHEET_CANDIDATES)
    sip_ws = _find_sheet(wb, SIP_SHEET_CANDIDATES)

    if holdings_ws is None:
        warnings.append("No holdings sheet found (looked for 'Mutual Fund' / similar).")
    if recs_ws is None:
        warnings.append("No recommendations sheet found (looked for 'Long Term' / 'Transaction Snapshot').")
    if sip_ws is None:
        warnings.append("No SIP sheet found (looked for 'SIP' / 'Sheet1').")

    holdings = _parse_holdings_or_recommendations_sheet(holdings_ws, warnings, "Mutual Fund")
    recommendations = _parse_holdings_or_recommendations_sheet(recs_ws, warnings, "Long Term")
    sips = _parse_sip_sheet(sip_ws, warnings)

    all_holdings = holdings + recommendations

    # Filter out any total rows that slipped through (belt & braces - the
    # per-sheet parser already skips them, this guards future call sites).
    all_holdings = [h for h in all_holdings if not _is_total_row(h.scheme)]

    wb.close()

    # Group into per-member results.
    members_by_name: dict[str, MemberResult] = {}

    def get_member(name: str, pan: Optional[str]) -> MemberResult:
        if name not in members_by_name:
            members_by_name[name] = MemberResult(name=name, pan=pan)
        elif pan and not members_by_name[name].pan:
            members_by_name[name].pan = pan
        return members_by_name[name]

    for h in all_holdings:
        m = get_member(h.member, h.pan)
        m.holdings.append(h)
        if h.purchase_value is not None:
            m.total_purchase_value += h.purchase_value
        if h.current_value is not None:
            m.total_current_value += h.current_value
        if h.gain is not None:
            m.total_gain += h.gain
        elif h.purchase_value is not None and h.current_value is not None:
            m.total_gain += h.current_value - h.purchase_value

    for s in sips:
        m = get_member(s.member, s.pan)
        m.sips.append(s)

    members = list(members_by_name.values())

    consolidated_purchase_value = sum(m.total_purchase_value for m in members)
    consolidated_current_value = sum(m.total_current_value for m in members)
    consolidated_gain = sum(m.total_gain for m in members)

    return PortfolioParseResult(
        members=members,
        holdings=all_holdings,
        sips=sips,
        consolidated_purchase_value=consolidated_purchase_value,
        consolidated_current_value=consolidated_current_value,
        consolidated_gain=consolidated_gain,
        warnings=warnings,
    )


# --------------------------------------------------------------------------
# Portfolio-level CAGR
# --------------------------------------------------------------------------

def compute_value_weighted_cagr(holdings: list[Holding]) -> Optional[float]:
    """Value-weighted average of each holding's own CAGR %, weighted by
    current value - e.g. a Rs 10L position at 15% CAGR moves the portfolio
    number more than a Rs 50K position at 40% CAGR.

    This is NOT a true portfolio XIRR (which needs per-lot cash-flow
    dates - purchase date per unit lot, SIP installment dates, etc. - that
    this pipeline doesn't currently track on Holding). It's a standard,
    defensible approximation, and it's actually COMPUTED from the
    holdings every call (unlike a hardcoded/caller-supplied figure, which
    is what this replaced - a static "12.8%" that never moved even when
    total invested and the holdings list changed).

    Holdings missing cagr_pct or current_value are excluded from both the
    numerator and denominator - never treated as a zero-CAGR holding,
    which would silently drag the average down.

    Returns None if no holding has both current_value and cagr_pct.
    """
    weighted_sum = 0.0
    weight_total = 0.0
    for h in holdings:
        if h.cagr_pct is None or h.current_value is None:
            continue
        weighted_sum += h.cagr_pct * h.current_value
        weight_total += h.current_value
    if weight_total <= 0:
        return None
    return weighted_sum / weight_total


# --------------------------------------------------------------------------
# Self-test (dummy in-memory workbook, no external files required)
# --------------------------------------------------------------------------

def _build_dummy_workbook(tmp_path: Path) -> Path:
    """Builds a small multi-member workbook exercising:
    - state-machine member switching (2 members)
    - embedded category header rows
    - a totals row that must be filtered out
    - one row with a missing current value (-> warning, not a guess)
    - a SIP sheet with a stop instruction
    """
    wb = openpyxl.Workbook()

    # --- Mutual Fund (holdings) sheet ---
    ws = wb.active
    ws.title = "Mutual Fund"
    rows = [
        ["Name: Rahul Sharma", None, None, None, None, None, None],
        ["PAN: ABCDE1234F", None, None, None, None, None, None],
        ["Equity", None, None, None, None, None, None],
        ["Scheme Name", "Folio No", "Balance Units", "Purchase Value", "Current Value", "Gain", "CAGR"],
        ["Axis Bluechip Fund", "F001", 1000.0, 50000.0, 62000.0, 12000.0, 14.5],
        ["HDFC Flexicap Fund", "F002", 500.0, 25000.0, None, None, None],  # missing current value
        ["Sub Total", None, None, 75000.0, 62000.0, 12000.0, None],
        ["Debt", None, None, None, None, None, None],
        ["ICICI Pru Short Term Fund", "F003", 200.0, 20000.0, 21500.0, 1500.0, 6.2],
        ["Total", None, None, 20000.0, 21500.0, 1500.0, None],
        ["Name: Priya Sharma", None, None, None, None, None, None],
        ["PAN: FGHIJ5678K", None, None, None, None, None, None],
        ["Equity", None, None, None, None, None, None],
        ["Scheme Name", "Folio No", "Balance Units", "Purchase Value", "Current Value", "Gain", "CAGR"],
        ["Mirae Asset Large Cap Fund", "F010", 800.0, 40000.0, 47000.0, 7000.0, 11.1],
        ["Grand Total", None, None, 40000.0, 47000.0, 7000.0, None],
    ]
    for r in rows:
        ws.append(r)

    # --- Long Term (recommendations) sheet ---
    ws2 = wb.create_sheet("Long Term")
    rows2 = [
        ["Name: Rahul Sharma", None, None, None],
        ["PAN: ABCDE1234F", None, None, None],
        ["Scheme Name", "Action", "Suggested Scheme", "Current Value"],
        ["HDFC Flexicap Fund", "Switch", "Parag Parikh Flexicap Fund", 26000.0],
    ]
    for r in rows2:
        ws2.append(r)

    # --- SIP sheet ---
    ws3 = wb.create_sheet("SIP")
    rows3 = [
        ["Name: Rahul Sharma", None, None, None],
        ["PAN: ABCDE1234F", None, None, None],
        ["Scheme Name", "SIP Amount", "SIP Date", "Instruction"],
        ["Axis Bluechip Fund", 5000.0, "5th", "Continue"],
        ["HDFC Flexicap Fund", 3000.0, "10th", "Stop"],
        ["Name: Priya Sharma", None, None, None],
        ["PAN: FGHIJ5678K", None, None, None],
        ["Scheme Name", "SIP Amount", "SIP Date", "Instruction"],
        ["Mirae Asset Large Cap Fund", None, "15th", "Start"],  # missing amount -> warning
    ]
    for r in rows3:
        ws3.append(r)

    out_path = tmp_path / "dummy_portfolio.xlsx"
    wb.save(out_path)
    wb.close()
    return out_path


def _run_self_test() -> None:
    import tempfile

    with tempfile.TemporaryDirectory() as tmp:
        dummy_path = _build_dummy_workbook(Path(tmp))
        result = parse_portfolio_workbook(dummy_path)

        print("=== pipeline/parser.py self-test ===")
        print(f"Members found: {[m.name for m in result.members]}")
        for m in result.members:
            print(
                f"  - {m.name} (PAN={m.pan}): "
                f"{len(m.holdings)} holdings, {len(m.sips)} SIPs, "
                f"purchase={m.total_purchase_value:.2f}, current={m.total_current_value:.2f}, "
                f"gain={m.total_gain:.2f}"
            )
        print(f"Total holdings (flat, all members): {len(result.holdings)}")
        for h in result.holdings:
            print(
                f"    [{h.member}] {h.scheme} | cat={h.category} | "
                f"purchase={h.purchase_value} current={h.current_value} "
                f"action={h.action} suggested={h.suggested_scheme}"
            )
        print(f"Total SIPs: {len(result.sips)}")
        for s in result.sips:
            print(f"    [{s.member}] {s.scheme} amount={s.sip_amount} instruction={s.instruction}")

        print(
            f"Consolidated: purchase={result.consolidated_purchase_value:.2f}, "
            f"current={result.consolidated_current_value:.2f}, gain={result.consolidated_gain:.2f}"
        )

        print(f"Warnings ({len(result.warnings)}):")
        for w in result.warnings:
            print(f"    - {w}")

        # Sanity assertions - fail loudly if totals rows leaked through.
        scheme_names_lower = {(h.scheme or "").strip().lower() for h in result.holdings}
        assert TOTAL_ROW_NAMES.isdisjoint(scheme_names_lower), "A totals row leaked into holdings!"
        assert len(result.members) == 2, f"Expected 2 members, got {len(result.members)}"
        assert any("current value missing" in w.lower() for w in result.warnings), (
            "Expected a warning for the missing current value row."
        )

        # --- compute_value_weighted_cagr ---
        print("\n--- compute_value_weighted_cagr ---")
        cagr_holdings = [
            Holding(member="X", pan=None, category="Equity", scheme="A", folio="1",
                    balance_units=None, purchase_value=None, current_value=1_000_000,
                    gain=None, absolute_return_pct=None, cagr_pct=15.0),
            Holding(member="X", pan=None, category="Equity", scheme="B", folio="2",
                    balance_units=None, purchase_value=None, current_value=50_000,
                    gain=None, absolute_return_pct=None, cagr_pct=40.0),
            Holding(member="X", pan=None, category="Debt", scheme="C", folio="3",
                    balance_units=None, purchase_value=None, current_value=200_000,
                    gain=None, absolute_return_pct=None, cagr_pct=None),  # excluded: no cagr_pct
        ]
        weighted_cagr = compute_value_weighted_cagr(cagr_holdings)
        expected = (1_000_000 * 15.0 + 50_000 * 40.0) / (1_000_000 + 50_000)
        print(f"Weighted CAGR: {weighted_cagr:.4f}%  (expected {expected:.4f}%)")
        assert weighted_cagr is not None and abs(weighted_cagr - expected) < 0.0001
        assert compute_value_weighted_cagr([]) is None

        print("\nAll self-test assertions passed.")


if __name__ == "__main__":
    _run_self_test()
