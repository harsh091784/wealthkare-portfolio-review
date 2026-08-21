"""
pipeline/ngen_lookup.py

NGEN returns matcher for the WC Securities / Wealthkare Portfolio Review
pipeline.

Given a fund name pulled from a client's portfolio, finds that fund's
trailing/CY return figures inside an NGEN market-stats export
(regular_ngen_ngen_markets_stats_[DATE].xlsx) and returns them as
percentages.

Design notes
------------
- NGEN exports are NOT keyed by a clean fund-code column we can rely on, so
  matching is done by fuzzy name comparison (Jaccard similarity over token
  sets, after stripping filler words like "fund"/"growth"/"direct"). Naive
  substring/regex matching fails on cases like "WhiteOak Capital" vs
  "WhiteOak" (regex needs an exact anchor; the token-set approach doesn't).
- Sheets are scanned in a fixed priority order (Equity, Hybrid, Debt, Other,
  Solution Oriented) with two documented special cases:
    * HDFC Balanced Advantage Fund lives in Hybrid, not Equity.
    * FOF category funds live in Other.
  Both are handled by simply checking Hybrid/Other early enough in the
  search order / with a dedicated pre-check, so no special-case branching
  is needed deep in the matching logic itself.
- Column layout inside each sheet is POSITIONALLY FIXED (this is an NGEN
  export quirk, not a WC Securities convention) - header text is not
  trustworthy, so we always index by column number, never by header lookup.
- If nothing matches above the similarity threshold, or the fund has no
  track record yet (blank cells), the field-level result is "N/A" - we
  never raise and never guess.
"""

from __future__ import annotations

import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Optional, Union

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

# --------------------------------------------------------------------------
# Constants
# --------------------------------------------------------------------------

# Sheet scan order. HDFC Balanced Advantage Fund is a documented exception
# that lives in Hybrid despite being an equity-sounding name; since Hybrid
# is scanned right after Equity, a caller doing a plain top-to-bottom scan
# will still find it without special-casing the matcher itself.
SHEET_SCAN_ORDER = ["Equity", "Hybrid", "Debt", "Other", "Solution Oriented"]

# Fund names that are documented to live in a sheet other than the one
# their category name would suggest. Checked first, before falling back to
# the general scan order, so we don't waste passes on wrong sheets.
KNOWN_SHEET_OVERRIDES: dict[str, str] = {
    "hdfc balanced advantage fund": "Hybrid",
}

# FOF-category funds live in the "Other" sheet. This is informational for
# callers that already know a fund is FOF-category; the general scan order
# also reaches "Other" on its own.
FOF_SHEET = "Other"

HEADER_ROW_INDEX = 4      # 0-based -> Excel row 5
DATA_START_ROW_INDEX = 5  # 0-based -> Excel row 6

# Positionally-fixed column indices (0-based) for each return period.
RETURN_COLUMNS: dict[str, int] = {
    "1Y": 15,
    "2Y": 16,
    "3Y": 17,
    "5Y": 18,
    "7Y": 19,
    "10Y": 20,
    "Since Launch": 22,
    "CY": 23,
    "CY-1": 24,
    "CY-2": 25,
    "CY-3": 26,
    "CY-4": 27,
}

NAME_COLUMNS = [0, 1, 2]  # scanned jointly as one combined string

STOPWORDS = {
    "fund", "reg", "regular", "capital", "the", "g", "plan", "growth",
    "direct", "option",
}

JACCARD_THRESHOLD = 0.55

NGEN_FILENAME_REGEX = re.compile(
    r"regular_ngen_ngen_markets_stats_(?P<date>[^./\\]+)\.xlsx$", re.IGNORECASE
)


# --------------------------------------------------------------------------
# Data model
# --------------------------------------------------------------------------

@dataclass
class NgenMatch:
    query_name: str
    matched_name: Optional[str] = None
    sheet: Optional[str] = None
    row: Optional[int] = None
    similarity: Optional[float] = None
    returns: dict[str, Union[float, str]] = field(default_factory=dict)  # "N/A" or pct float
    found: bool = False
    warnings: list[str] = field(default_factory=list)


# --------------------------------------------------------------------------
# Jaccard matching
# --------------------------------------------------------------------------

def jaccard_similarity(name_a: str, name_b: str, stopwords: set[str]) -> float:
    tokens_a = set(name_a.lower().split()) - stopwords
    tokens_b = set(name_b.lower().split()) - stopwords
    if not tokens_a or not tokens_b:
        return 0.0
    return len(tokens_a & tokens_b) / len(tokens_a | tokens_b)


def _tokenize_for_compare(text: str) -> str:
    """Light normalization before tokenizing: lowercase and split words on
    non-alphanumerics so punctuation doesn't create spurious tokens."""
    return " ".join(re.findall(r"[a-z0-9]+", text.lower()))


def _combined_name_from_row(values: list) -> str:
    parts = []
    for idx in NAME_COLUMNS:
        if idx < len(values) and values[idx] is not None:
            text = str(values[idx]).strip()
            if text:
                parts.append(text)
    return " ".join(parts)


# --------------------------------------------------------------------------
# Sheet scanning
# --------------------------------------------------------------------------

def _best_match_in_sheet(
    ws: Worksheet, query_name: str, threshold: float
) -> tuple[Optional[str], Optional[int], float]:
    """Returns (matched_combined_name, 0-based row index, similarity) for
    the best-scoring row in this sheet, or (None, None, 0.0) if nothing
    clears the threshold."""
    best_name: Optional[str] = None
    best_row: Optional[int] = None
    best_score = 0.0

    query_norm = _tokenize_for_compare(query_name)

    max_col = ws.max_column or (max(RETURN_COLUMNS.values()) + 1)
    max_row = ws.max_row or 0
    if max_row <= DATA_START_ROW_INDEX:
        return None, None, 0.0

    for row_idx_0based in range(DATA_START_ROW_INDEX, max_row):
        excel_row = row_idx_0based + 1  # openpyxl rows are 1-based
        values = [ws.cell(row=excel_row, column=c + 1).value for c in range(max_col)]
        combined = _combined_name_from_row(values)
        if not combined:
            continue
        combined_norm = _tokenize_for_compare(combined)
        score = jaccard_similarity(query_norm, combined_norm, STOPWORDS)
        if score > best_score:
            best_score = score
            best_name = combined
            best_row = row_idx_0based

    if best_score >= threshold:
        return best_name, best_row, best_score
    return None, None, best_score


def _extract_returns(ws: Worksheet, row_idx_0based: int) -> dict[str, Union[float, str]]:
    excel_row = row_idx_0based + 1
    returns: dict[str, Union[float, str]] = {}
    for label, col_idx in RETURN_COLUMNS.items():
        raw = ws.cell(row=excel_row, column=col_idx + 1).value
        if raw is None or (isinstance(raw, str) and raw.strip() in ("", "-", "NA", "N/A")):
            returns[label] = "N/A"
            continue
        try:
            pct = float(raw) * 100
            returns[label] = round(pct, 2)
        except (TypeError, ValueError):
            returns[label] = "N/A"
    return returns


# --------------------------------------------------------------------------
# Top-level entry point
# --------------------------------------------------------------------------

def find_fund_returns(
    workbook_path: Union[str, Path],
    fund_name: str,
    category_hint: Optional[str] = None,
    threshold: float = JACCARD_THRESHOLD,
) -> NgenMatch:
    """Look up `fund_name` inside an NGEN market-stats workbook and return
    its trailing/CY returns as percentages.

    category_hint: optional category ("Equity"/"Hybrid"/"Debt"/"Other"/
    "Solution Oriented"/"FOF") to check first, before falling back to the
    documented overrides and the general scan order. Never required -
    matching works without it, but supplying it (when known) saves scan
    passes and reduces the chance of a false positive in the wrong sheet.
    """
    workbook_path = Path(workbook_path)
    result = NgenMatch(query_name=fund_name)

    if not NGEN_FILENAME_REGEX.search(workbook_path.name):
        result.warnings.append(
            f"Filename '{workbook_path.name}' doesn't match the expected "
            f"'regular_ngen_ngen_markets_stats_[DATE].xlsx' pattern - proceeding anyway."
        )

    wb = openpyxl.load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        sheet_names = set(wb.sheetnames)

        # Build the ordered list of sheets to try: known override first (if
        # this exact fund name has one), then an optional caller-supplied
        # category hint, then the documented default scan order.
        candidate_sheets: list[str] = []

        override_sheet = KNOWN_SHEET_OVERRIDES.get(fund_name.strip().lower())
        if override_sheet:
            candidate_sheets.append(override_sheet)

        if category_hint:
            normalized_hint = category_hint.strip().lower()
            if normalized_hint == "fof":
                candidate_sheets.append(FOF_SHEET)
            else:
                for sheet_name in SHEET_SCAN_ORDER:
                    if sheet_name.lower() == normalized_hint:
                        candidate_sheets.append(sheet_name)
                        break

        for sheet_name in SHEET_SCAN_ORDER:
            if sheet_name not in candidate_sheets:
                candidate_sheets.append(sheet_name)

        best_overall_name = None
        best_overall_row = None
        best_overall_score = 0.0
        best_overall_sheet = None

        for sheet_name in candidate_sheets:
            if sheet_name not in sheet_names:
                continue
            ws = wb[sheet_name]
            matched_name, row_idx, score = _best_match_in_sheet(ws, fund_name, threshold)
            if matched_name is not None and score > best_overall_score:
                best_overall_name = matched_name
                best_overall_row = row_idx
                best_overall_score = score
                best_overall_sheet = sheet_name
                # Good enough match found - no need to keep scanning further
                # sheets once we've cleared the threshold in priority order.
                break

        if best_overall_name is None:
            result.found = False
            result.similarity = best_overall_score
            result.returns = {label: "N/A" for label in RETURN_COLUMNS}
            result.warnings.append(
                f"No fund matched '{fund_name}' above similarity threshold {threshold} "
                f"in any sheet ({', '.join(candidate_sheets)})."
            )
            return result

        ws = wb[best_overall_sheet]
        returns = _extract_returns(ws, best_overall_row)

        # If every return field came back N/A, treat this as "no track
        # record yet" (e.g. newly launched fund) rather than a hard error.
        if all(v == "N/A" for v in returns.values()):
            result.warnings.append(
                f"Matched '{fund_name}' -> '{best_overall_name}' in sheet "
                f"'{best_overall_sheet}' but it has no track record data (newly launched?)."
            )

        result.matched_name = best_overall_name
        result.sheet = best_overall_sheet
        result.row = best_overall_row + 1  # report as 1-based Excel row
        result.similarity = round(best_overall_score, 3)
        result.returns = returns
        result.found = True
        return result
    finally:
        wb.close()


# --------------------------------------------------------------------------
# Self-test (dummy in-memory workbook, no external files required)
# --------------------------------------------------------------------------

def _build_dummy_ngen_workbook(tmp_path: Path) -> Path:
    """Builds a small NGEN-shaped workbook:
    - header row at Excel row 5 (index 4), data starting row 6 (index 5)
    - 28 columns so all fixed positional indices (up to 27) are populated
    - one Equity fund, one Hybrid fund (HDFC Balanced Advantage - the
      documented override case), one Other/FOF fund, and one fund with a
      blank track record (newly launched) to exercise the N/A path.
    """
    wb = openpyxl.Workbook()

    def make_header(ws: Worksheet) -> None:
        # Rows 1-4 are whatever preamble NGEN puts there; row 5 (index 4) is
        # the header row. We only need row 5+ to have the right SHAPE (28
        # columns) - header text itself is deliberately not meaningful,
        # since column lookup is positional, not header-driven.
        for r in range(1, 5):
            ws.append([None] * 28)
        header = [f"col{i}" for i in range(28)]
        header[0], header[1], header[2] = "AMC", "Scheme", "Plan"
        ws.append(header)

    def make_return_row(scheme_combo: tuple, returns_by_col: dict) -> list:
        row = [None] * 28
        row[0], row[1], row[2] = scheme_combo
        for col_idx, val in returns_by_col.items():
            row[col_idx] = val
        return row

    sample_returns = {
        15: 0.1428,  # 1Y = 14.28%
        16: 0.1120,  # 2Y
        17: 0.1350,  # 3Y
        18: 0.1210,  # 5Y
        19: 0.1180,  # 7Y
        20: 0.1155,  # 10Y
        22: 0.1300,  # Since Launch
        23: 0.0950,  # CY
        24: 0.1400,  # CY-1
        25: 0.0800,  # CY-2
        26: 0.2200,  # CY-3
        27: 0.0500,  # CY-4
    }

    # --- Equity sheet ---
    ws_eq = wb.active
    ws_eq.title = "Equity"
    make_header(ws_eq)
    ws_eq.append(make_return_row(("WhiteOak Capital", "Flexi Cap Fund", "Direct Growth"), sample_returns))
    ws_eq.append(make_return_row(("Axis", "Bluechip Fund", "Reg Growth"), sample_returns))
    # newly launched fund - all return cells blank -> should resolve to N/A everywhere
    ws_eq.append(make_return_row(("Newco", "Momentum Fund", "Direct Growth"), {}))

    # --- Hybrid sheet: HDFC Balanced Advantage override case ---
    ws_hy = wb.create_sheet("Hybrid")
    make_header(ws_hy)
    ws_hy.append(make_return_row(("HDFC", "Balanced Advantage Fund", "Direct Growth"), sample_returns))

    # --- Debt sheet ---
    ws_debt = wb.create_sheet("Debt")
    make_header(ws_debt)
    ws_debt.append(make_return_row(("ICICI Prudential", "Short Term Fund", "Direct Growth"), sample_returns))

    # --- Other sheet: FOF example ---
    ws_other = wb.create_sheet("Other")
    make_header(ws_other)
    ws_other.append(make_return_row(("Motilal Oswal", "Nasdaq 100 FOF", "Direct Growth"), sample_returns))

    # --- Solution Oriented sheet (empty but present) ---
    wb.create_sheet("Solution Oriented")

    out_path = tmp_path / "regular_ngen_ngen_markets_stats_31Jul2026.xlsx"
    wb.save(out_path)
    wb.close()
    return out_path


def _run_self_test() -> None:
    import tempfile

    with tempfile.TemporaryDirectory() as tmp:
        dummy_path = _build_dummy_ngen_workbook(Path(tmp))

        print("=== pipeline/ngen_lookup.py self-test ===")

        cases = [
            # (query name, category hint, expected to be found)
            ("WhiteOak Capital Flexi Cap Fund", None, True),   # substring-hostile name, should match via tokens
            ("HDFC Balanced Advantage Fund", None, True),      # documented Hybrid override
            ("Motilal Oswal Nasdaq 100 FOF", "FOF", True),     # FOF -> Other sheet
            ("Axis Bluechip Fund", "Equity", True),
            ("Totally Unknown Fund XYZ", None, False),         # should stay unmatched -> N/A
            ("Newco Momentum Fund", None, True),               # matches, but no track record -> N/A returns
        ]

        for query, hint, expect_found in cases:
            result = find_fund_returns(dummy_path, query, category_hint=hint)
            status = "FOUND" if result.found else "NOT FOUND"
            print(f"\nQuery: '{query}' (hint={hint}) -> {status}")
            if result.found:
                print(f"  Matched: '{result.matched_name}' in sheet '{result.sheet}' (row {result.row}, similarity={result.similarity})")
                print(f"  Returns: {result.returns}")
            for w in result.warnings:
                print(f"  Warning: {w}")

            assert result.found == expect_found, f"Expected found={expect_found} for '{query}', got {result.found}"

        # Spot-check actual percentage conversion on a known-good match.
        wo_result = find_fund_returns(dummy_path, "WhiteOak Capital Flexi Cap Fund")
        assert wo_result.returns["1Y"] == 14.28, f"Expected 1Y=14.28, got {wo_result.returns['1Y']}"
        assert wo_result.returns["CY-3"] == 22.0, f"Expected CY-3=22.0, got {wo_result.returns['CY-3']}"

        # HDFC Balanced Advantage must resolve via the Hybrid sheet, not Equity.
        hdfc_result = find_fund_returns(dummy_path, "HDFC Balanced Advantage Fund")
        assert hdfc_result.sheet == "Hybrid", f"Expected HDFC Balanced Advantage in Hybrid sheet, got {hdfc_result.sheet}"

        # Newly launched fund -> matched, but every return field is N/A.
        newco_result = find_fund_returns(dummy_path, "Newco Momentum Fund")
        assert all(v == "N/A" for v in newco_result.returns.values()), "Expected all N/A for newly launched fund."

        # Unknown fund -> not found, returns dict is all N/A, no exception raised.
        unknown_result = find_fund_returns(dummy_path, "Totally Unknown Fund XYZ")
        assert all(v == "N/A" for v in unknown_result.returns.values()), "Expected all N/A for unmatched fund."

        # jaccard_similarity direct sanity check on the exact example from spec.
        score = jaccard_similarity("whiteoak capital flexi cap fund", "whiteoak flexi cap direct growth", STOPWORDS)
        print(f"\nDirect jaccard_similarity('WhiteOak Capital...' vs 'WhiteOak Flexi Cap...') = {score:.3f}")
        assert score >= JACCARD_THRESHOLD, "Expected WhiteOak variants to clear the similarity threshold."

        print("\nAll self-test assertions passed.")


if __name__ == "__main__":
    _run_self_test()
