"""
tests/test_dashboard_parser.py

Every documented quirk of the real dashboard export, reproduced as a
synthetic workbook.

WHY SYNTHETIC: the four real client files are not in this repo, so these
tests encode the SPECIFIED behaviour rather than confirming it against
those files. Each test names the real-file observation it stands in for.
When the real files land, run them through and these become the
regression net rather than the specification.
"""

from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import Workbook

from pipeline.dashboard_parser import (
    ASSET_CLASS_ROWS,
    MAX_PLAUSIBLE_CAGR_PCT,
    RECONCILE_TOLERANCE_RUPEES,
    ParseError,
    canonical_header,
    classify_action,
    normalise_header,
    parse_dashboard_workbook,
    split_client_label,
)

FULL_HEADERS = [
    "Scheme", "Folio No", "Balance Units", "Purchase Value", "Current Value",
    "Gain", "Holding Days", "Absolute Return (%)", "CAGR (%)",
]


def _holding_row(scheme, folio, units, purchase, current, gain, days, abs_ret, cagr, extra=None):
    return [scheme, folio, units, purchase, current, gain, days, abs_ret, cagr] + (extra or [])


def _write(tmp_path, sheets: dict, name="upload.xlsx"):
    """sheets: {sheet_name: [row, row, ...]} in insertion order."""
    wb = Workbook()
    wb.remove(wb.active)
    for sheet_name, rows in sheets.items():
        ws = wb.create_sheet(title=sheet_name)
        for row in rows:
            ws.append(row)
    path = tmp_path / name
    wb.save(path)
    return path


def _simple_book(tmp_path, holdings_sheet_name="1. Mutual Fund", **kwargs):
    rows = [
        ["Portfolio Holdings Report"], [],
        FULL_HEADERS,
        ["PRIYA SHARMA (AAAAA1111A)"],
        ["Equity"],
        _holding_row("Alpha Large Cap Fund", "1001", 100.0, 100000, 130000, 30000, 800, 30.0, 12.5),
        _holding_row("Beta Mid Cap Fund", "1002", 200.0, 200000, 260000, 60000, 900, 30.0, 13.0),
        ["Grand Total", None, None, 300000, 390000, 90000],
    ]
    return _write(tmp_path, {holdings_sheet_name: rows}, **kwargs)


# --------------------------------------------------------------------------
# Header handling
# --------------------------------------------------------------------------

def test_header_normalisation_strips_collapses_and_uppercases():
    assert normalise_header("  folio   no ") == "FOLIO NO"
    assert normalise_header("CAGR (%)") == "CAGR (%)"
    assert normalise_header(None) == ""


@pytest.mark.parametrize("raw,expected", [
    ("Suggestion", "SUGGESTED SCHEME"),
    ("  actual   capital gain ", "GAIN"),
    ("Purchase Amount", "PURCHASE VALUE"),
    ("Scheme", "SCHEME"),
])
def test_header_synonyms_map_to_canonical_names(raw, expected):
    assert canonical_header(raw) == expected


def test_columns_are_located_by_name_not_position(tmp_path):
    """The whole point of header-based parsing: shuffle the column order
    and the same values must come back attached to the same fields."""
    shuffled = ["CAGR (%)", "Gain", "Scheme", "Current Value", "Absolute Return (%)",
                "Folio No", "Holding Days", "Purchase Value", "Balance Units"]
    rows = [
        shuffled,
        ["ROHAN IYER (ABCDE1234F)"],
        [12.5, 30000, "Alpha Large Cap Fund", 130000, 30.0, "1001", 800, 100000, 100.0],
    ]
    result = parse_dashboard_workbook(_write(tmp_path, {"1. Mutual Fund": rows}))
    holding = result.clients[0].holdings[0]
    assert holding.scheme == "Alpha Large Cap Fund"
    assert holding.current_value == 130000
    assert holding.purchase_value == 100000
    assert holding.cagr_pct == 12.5
    assert holding.folio == "1001"


def test_missing_required_column_fails_loudly_naming_sheet_and_columns(tmp_path):
    headers = [h for h in FULL_HEADERS if h != "CAGR (%)"]
    rows = [headers, ["PRIYA SHARMA (AAAAA1111A)"],
            ["Alpha Large Cap Fund", "1001", 100.0, 100000, 130000, 30000, 800, 30.0]]
    with pytest.raises(ParseError) as excinfo:
        parse_dashboard_workbook(_write(tmp_path, {"1. Mutual Fund": rows}))
    message = str(excinfo.value)
    assert "CAGR (%)" in message
    assert "1. Mutual Fund" in message, "the error must name the sheet"


# --------------------------------------------------------------------------
# Sheet detection - real files number their tabs inconsistently
# --------------------------------------------------------------------------

@pytest.mark.parametrize("sheet_name", ["1. Mutual Fund", "Mutual Fund", "MUTUAL FUND", "2. mutual fund"])
def test_holdings_sheet_found_by_case_insensitive_substring(sheet_name, tmp_path):
    result = parse_dashboard_workbook(_simple_book(tmp_path, holdings_sheet_name=sheet_name))
    assert len(result.clients) == 1
    assert len(result.clients[0].holdings) == 2


@pytest.mark.parametrize("actions_name", ["Long Term", "2. Long Term", "LONG TERM"])
def test_actions_sheet_found_by_case_insensitive_substring(actions_name, tmp_path):
    holdings = [
        FULL_HEADERS, ["PRIYA SHARMA (AAAAA1111A)"],
        _holding_row("Alpha Large Cap Fund", "1001", 100.0, 100000, 130000, 30000, 800, 30.0, 12.5),
    ]
    actions = [["Scheme", "Folio No", "Action", "Suggested Scheme"],
               ["Alpha Large Cap Fund", "1001", "Switch", "Gamma Flexi Cap Fund"]]
    result = parse_dashboard_workbook(
        _write(tmp_path, {"1. Mutual Fund": holdings, actions_name: actions})
    )
    assert [a.canonical for a in result.clients[0].actions] == ["switch"]


@pytest.mark.parametrize("sip_name", ["SIP", "Sip", "3. SIP"])
def test_sip_sheet_found_by_case_insensitive_substring(sip_name, tmp_path):
    holdings = [
        FULL_HEADERS, ["PRIYA SHARMA (AAAAA1111A)"],
        _holding_row("Alpha Large Cap Fund", "1001", 100.0, 100000, 130000, 30000, 800, 30.0, 12.5),
    ]
    sips = [["Scheme", "SIP Amount"], ["Alpha Large Cap Fund", 5000]]
    result = parse_dashboard_workbook(_write(tmp_path, {"1. Mutual Fund": holdings, sip_name: sips}))
    assert len(result.clients[0].sips) == 1
    assert result.clients[0].sips[0].sip_amount == 5000


def test_missing_holdings_sheet_names_what_the_file_actually_has(tmp_path):
    with pytest.raises(ParseError) as excinfo:
        parse_dashboard_workbook(_write(tmp_path, {"Summary": [["nothing here"]]}))
    assert "Summary" in str(excinfo.value)


# --------------------------------------------------------------------------
# Client vs asset-class header rows
# --------------------------------------------------------------------------

def test_clients_and_asset_classes_are_told_apart_by_the_known_set(tmp_path):
    """Both are "text in column A, everything else empty". The asset-class
    set is what separates them - NOT a PAN-shaped test."""
    rows = [FULL_HEADERS,
            ["PRIYA SHARMA (AAAAA1111A)"], ["Equity"],
            _holding_row("Alpha Large Cap Fund", "1001", 100.0, 100000, 130000, 30000, 800, 30.0, 12.5),
            ["Debt"],
            _holding_row("Delta Short Term Fund", "1003", 50.0, 50000, 53000, 3000, 400, 6.0, 5.5),
            ["RAJESH KUMAR (BCDEF2345G)"], ["Hybrid"],
            _holding_row("Epsilon BAF", "2001", 80.0, 80000, 92000, 12000, 600, 15.0, 9.0)]
    result = parse_dashboard_workbook(_write(tmp_path, {"1. Mutual Fund": rows}))

    assert result.client_names() == ["PRIYA SHARMA", "RAJESH KUMAR"]
    sonali = result.client("PRIYA SHARMA")
    assert [h.category for h in sonali.holdings] == ["Equity", "Debt"]
    assert [h.category for h in result.client("RAJESH KUMAR").holdings] == ["Hybrid"]


def test_company_client_with_truncated_pan_is_not_mistaken_for_an_asset_class(tmp_path):
    """One real export carries a corporate client whose bracketed PAN is
    truncated. A PAN-pattern test would fail to recognise it as a client
    and file its holdings under whoever came before - which is why client
    rows are told apart from asset-class rows by the known asset-class
    set, not by matching a PAN. Modelled here with invented values."""
    rows = [FULL_HEADERS,
            ["PRIYA SHARMA (AAAAA1111A)"], ["Equity"],
            _holding_row("Alpha Large Cap Fund", "1001", 100.0, 100000, 130000, 30000, 800, 30.0, 12.5),
            ["EXAMPLE HOLDINGS PVT LTD (CCCCC33)"], ["Equity"],
            _holding_row("Zeta Value Fund", "3001", 10.0, 10000, 11000, 1000, 300, 10.0, 8.0)]
    result = parse_dashboard_workbook(_write(tmp_path, {"1. Mutual Fund": rows}))

    assert "EXAMPLE HOLDINGS PVT LTD" in result.client_names()
    company = result.client("EXAMPLE HOLDINGS PVT LTD")
    assert company.pan == "CCCCC33", "a truncated PAN is recorded as-is, not discarded"
    assert len(company.holdings) == 1
    assert len(result.client("PRIYA SHARMA").holdings) == 1, "holdings must not leak across clients"


def test_client_with_no_pan_brackets_is_still_a_client(tmp_path):
    rows = [FULL_HEADERS, ["EXAMPLE HOLDINGS PVT LTD"],
            _holding_row("Zeta Value Fund", "3001", 10.0, 10000, 11000, 1000, 300, 10.0, 8.0)]
    result = parse_dashboard_workbook(_write(tmp_path, {"1. Mutual Fund": rows}))
    assert result.client_names() == ["EXAMPLE HOLDINGS PVT LTD"]
    assert result.clients[0].pan is None


@pytest.mark.parametrize("asset_class", sorted(ASSET_CLASS_ROWS))
def test_every_asset_class_row_is_recognised_as_a_category(asset_class, tmp_path):
    rows = [FULL_HEADERS, ["PRIYA SHARMA (AAAAA1111A)"], [asset_class.title()],
            _holding_row("Some Fund", "1001", 10.0, 10000, 11000, 1000, 300, 10.0, 8.0)]
    result = parse_dashboard_workbook(_write(tmp_path, {"1. Mutual Fund": rows}))
    assert result.client_names() == ["PRIYA SHARMA"], f"{asset_class} was treated as a client"
    assert result.clients[0].holdings[0].category == asset_class.title()


def test_name_column_groups_clients_when_there_are_no_header_rows(tmp_path):
    rows = [["Name"] + FULL_HEADERS,
            ["PRIYA SHARMA (AAAAA1111A)"] + _holding_row("Alpha Large Cap Fund", "1001", 100.0, 100000, 130000, 30000, 800, 30.0, 12.5),
            ["PRIYA SHARMA (AAAAA1111A)"] + _holding_row("Beta Mid Cap Fund", "1002", 50.0, 50000, 60000, 10000, 700, 20.0, 10.0),
            ["RAJESH KUMAR (BCDEF2345G)"] + _holding_row("Epsilon BAF", "2001", 80.0, 80000, 92000, 12000, 600, 15.0, 9.0)]
    result = parse_dashboard_workbook(_write(tmp_path, {"1. Mutual Fund": rows}))
    assert result.client_names() == ["PRIYA SHARMA", "RAJESH KUMAR"]
    assert len(result.client("PRIYA SHARMA").holdings) == 2
    assert len(result.client("RAJESH KUMAR").holdings) == 1


# --------------------------------------------------------------------------
# The CAGR-shift defect
# --------------------------------------------------------------------------

def test_cagr_shift_is_corrected_by_reading_one_column_right(tmp_path):
    """The real observation: gain=185564.32, CAGR=185564.32, real=97.31."""
    rows = [FULL_HEADERS + ["Spill"],
            ["ROHAN IYER (ABCDE1234F)"],
            _holding_row("Shifted Fund", "1001", 100.0, 200000, 385564.32, 185564.32,
                         800, 92.78, 185564.32, extra=[97.31])]
    result = parse_dashboard_workbook(_write(tmp_path, {"1. Mutual Fund": rows}))
    client = result.clients[0]

    assert client.holdings[0].cagr_pct == 97.31, "the real CAGR one column right was not picked up"
    corrected = [w for w in client.warnings if w.kind == "cagr_shift_corrected"]
    assert len(corrected) == 1
    assert "97.31" in corrected[0].message and "Shifted Fund" in corrected[0].message


def test_absurd_cagr_with_no_usable_neighbour_is_rejected_not_ingested(tmp_path):
    """Never ingest a 4,129% CAGR. With nothing usable to the right the
    value is discarded and the row is kept, so the grand total still
    reconciles."""
    rows = [FULL_HEADERS + ["Spill"],
            ["ROHAN IYER (ABCDE1234F)"],
            _holding_row("Absurd Fund", "1001", 100.0, 100000, 130000, 30000,
                         800, 30.0, 4129.0, extra=[None])]
    result = parse_dashboard_workbook(_write(tmp_path, {"1. Mutual Fund": rows}))
    client = result.clients[0]

    assert client.holdings[0].cagr_pct is None
    assert client.holdings[0].current_value == 130000, "the rest of the row must survive"
    rejected = [w for w in client.warnings if w.kind == "cagr_rejected"]
    assert len(rejected) == 1 and "4,129" in rejected[0].message


def test_plausible_cagr_is_left_alone(tmp_path):
    result = parse_dashboard_workbook(_simple_book(tmp_path))
    assert [h.cagr_pct for h in result.clients[0].holdings] == [12.5, 13.0]
    assert not any(w.kind.startswith("cagr") for w in result.clients[0].warnings)


def test_cagr_equal_to_gain_is_detected_even_when_not_absurdly_large(tmp_path):
    """Both detectors matter: a small gain duplicated into CAGR produces a
    plausible-looking number that is still wrong."""
    rows = [FULL_HEADERS + ["Spill"], ["ROHAN IYER (ABCDE1234F)"],
            _holding_row("Quiet Shift Fund", "1001", 100.0, 100000, 100085, 85.0,
                         800, 0.085, 85.0, extra=[11.4])]
    result = parse_dashboard_workbook(_write(tmp_path, {"1. Mutual Fund": rows}))
    assert result.clients[0].holdings[0].cagr_pct == 11.4
    assert any(w.kind == "cagr_shift_corrected" for w in result.clients[0].warnings)


def test_max_plausible_cagr_threshold_is_below_the_observed_defect():
    assert MAX_PLAUSIBLE_CAGR_PCT < 4129.0


# --------------------------------------------------------------------------
# Actions
# --------------------------------------------------------------------------

@pytest.mark.parametrize("raw,kind", [
    ("Switch", "transaction"), ("REDEEM", "transaction"), ("trim", "transaction"),
    ("Watchlist", "things_to_do"), ("accumulate", "things_to_do"),
    ("Tax Loss Harvest", "things_to_do"), ("TAX  LOSS  HARVEST", "things_to_do"),
    ("swtich", "unrecognised"), ("hold", "unrecognised"), ("", "none"),
])
def test_action_vocabulary_is_matched_case_insensitively(raw, kind):
    assert classify_action(raw)[0] == kind


def test_unrecognised_action_becomes_a_warning_and_is_never_applied(tmp_path):
    holdings = [FULL_HEADERS, ["PRIYA SHARMA (AAAAA1111A)"],
                _holding_row("Alpha Large Cap Fund", "1001", 100.0, 100000, 130000, 30000, 800, 30.0, 12.5)]
    actions = [["Scheme", "Folio No", "Action"], ["Alpha Large Cap Fund", "1001", "Reduce Slightly"]]
    result = parse_dashboard_workbook(_write(tmp_path, {"1. Mutual Fund": holdings, "Long Term": actions}))
    client = result.clients[0]

    warned = [w for w in client.warnings if w.kind == "action_unrecognised"]
    assert len(warned) == 1, "an unknown action must surface, never be silently dropped"
    assert "Reduce Slightly" in warned[0].message
    assert client.holdings[0].action is None, "an unrecognised action must not be applied"
    assert [a.action_raw for a in client.actions] == ["Reduce Slightly"]


def test_action_matched_to_holding_by_scheme_and_folio(tmp_path):
    holdings = [FULL_HEADERS, ["PRIYA SHARMA (AAAAA1111A)"],
                _holding_row("Alpha Large Cap Fund", "1001", 100.0, 100000, 130000, 30000, 800, 30.0, 12.5),
                _holding_row("Alpha Large Cap Fund", "9999", 10.0, 10000, 11000, 1000, 300, 10.0, 8.0)]
    actions = [["Scheme", "Folio No", "Action", "Suggestion"],
               ["alpha large cap fund", "9999", "redeem", "Gamma Flexi Cap Fund"]]
    result = parse_dashboard_workbook(_write(tmp_path, {"1. Mutual Fund": holdings, "Long Term": actions}))
    by_folio = {h.folio: h for h in result.clients[0].holdings}

    assert by_folio["9999"].action == "redeem", "matching must be case-insensitive on scheme"
    assert by_folio["9999"].suggested_scheme == "Gamma Flexi Cap Fund"
    assert by_folio["1001"].action is None, "the wrong folio must not receive the action"


def test_unmatched_action_is_warned_not_dropped_silently(tmp_path):
    holdings = [FULL_HEADERS, ["PRIYA SHARMA (AAAAA1111A)"],
                _holding_row("Alpha Large Cap Fund", "1001", 100.0, 100000, 130000, 30000, 800, 30.0, 12.5)]
    actions = [["Scheme", "Folio No", "Action"], ["A Fund Nobody Holds", "7777", "Switch"]]
    result = parse_dashboard_workbook(_write(tmp_path, {"1. Mutual Fund": holdings, "Long Term": actions}))
    assert any(w.kind == "action_unmatched" for w in result.warnings)


def test_zero_actions_is_a_valid_file(tmp_path):
    """One real client has no actions at all. That must parse cleanly and
    produce an empty transaction set, not an error."""
    holdings = [FULL_HEADERS, ["JITENDER SINGH (CDEFG3456H)"],
                _holding_row("Alpha Large Cap Fund", "1001", 100.0, 100000, 130000, 30000, 800, 30.0, 12.5)]
    actions = [["Scheme", "Folio No", "Action"], ["Alpha Large Cap Fund", "1001", None]]
    result = parse_dashboard_workbook(_write(tmp_path, {"1. Mutual Fund": holdings, "Long Term": actions}))
    client = result.clients[0]

    assert client.actions == []
    assert len(client.holdings) == 1
    assert all(h.action is None for h in client.holdings)


def test_no_actions_sheet_at_all_still_parses(tmp_path):
    result = parse_dashboard_workbook(_simple_book(tmp_path))
    assert len(result.clients[0].holdings) == 2
    assert result.clients[0].actions == []
    assert any(w.kind == "no_actions_sheet" for w in result.warnings)


# --------------------------------------------------------------------------
# Grand Total reconciliation
# --------------------------------------------------------------------------

def test_grand_total_row_is_workbook_level_not_per_client(tmp_path):
    """"Grand Total" totals the whole sheet, not the last client on it.

    Every real export carries BOTH kinds of total row - a per-client
    "<NAME> Total" and one workbook-level "Grand Total" - and attaching
    the workbook figure to whichever client happened to be open last
    compared one client's holdings against the entire book. On MANISH's
    file that read as a Rs 28.16 crore discrepancy when nothing was
    wrong with the data at all.
    """
    result = parse_dashboard_workbook(_simple_book(tmp_path))
    assert result.reported_workbook_total == 390000
    assert result.computed_workbook_total == 390000
    assert result.workbook_total_reconciles is True
    # ...and it did NOT get pinned onto the client.
    assert result.clients[0].reported_grand_total is None


def test_per_client_total_row_reconciles_that_client(tmp_path):
    """A "<NAME> Total" row belongs to the client it names."""
    rows = [FULL_HEADERS,
            ["PRIYA SHARMA (AAAAA1111A)"],
            _holding_row("Alpha Large Cap Fund", "1001", 100.0, 100000, 130000, 30000, 800, 30.0, 12.5),
            ["PRIYA SHARMA Total", None, None, 100000, 130000],
            ["ARJUN MEHTA (BBBBB2222B)"],
            _holding_row("Beta Mid Cap Fund", "1002", 50.0, 200000, 260000, 60000, 800, 30.0, 12.5),
            ["ARJUN MEHTA Total", None, None, 200000, 260000],
            ["Grand Total", None, None, 300000, 390000]]
    result = parse_dashboard_workbook(_write(tmp_path, {"1. Mutual Fund": rows}))
    assert [c.name for c in result.clients] == ["PRIYA SHARMA", "ARJUN MEHTA"]
    assert result.clients[0].reported_grand_total == 130000
    assert result.clients[1].reported_grand_total == 260000
    assert all(c.grand_total_reconciles is True for c in result.clients)
    assert result.reported_workbook_total == 390000
    assert result.workbook_total_reconciles is True


def test_per_client_total_mismatch_is_reported(tmp_path):
    rows = [FULL_HEADERS, ["PRIYA SHARMA (AAAAA1111A)"],
            _holding_row("Alpha Large Cap Fund", "1001", 100.0, 100000, 130000, 30000, 800, 30.0, 12.5),
            ["PRIYA SHARMA Total", None, None, 100000, 999999]]
    result = parse_dashboard_workbook(_write(tmp_path, {"1. Mutual Fund": rows}))
    client = result.clients[0]
    assert client.grand_total_reconciles is False
    assert any(w.kind == "grand_total_mismatch" for w in client.warnings)


def test_workbook_total_mismatch_is_reported(tmp_path):
    """The book-level check is the one that catches a client being lost
    entirely: every per-client total can reconcile while the book is
    short by a whole client."""
    rows = [FULL_HEADERS, ["PRIYA SHARMA (AAAAA1111A)"],
            _holding_row("Alpha Large Cap Fund", "1001", 100.0, 100000, 130000, 30000, 800, 30.0, 12.5),
            ["PRIYA SHARMA Total", None, None, 100000, 130000],
            ["Grand Total", None, None, 300000, 999999]]
    result = parse_dashboard_workbook(_write(tmp_path, {"1. Mutual Fund": rows}))
    assert result.clients[0].grand_total_reconciles is True
    assert result.workbook_total_reconciles is False
    assert any(w.kind == "workbook_total_mismatch" for w in result.warnings)


def test_absent_grand_total_reports_unknown_not_pass(tmp_path):
    """No Grand Total row means the check could not be run. That is not
    the same as the check passing."""
    rows = [FULL_HEADERS, ["PRIYA SHARMA (AAAAA1111A)"],
            _holding_row("Alpha Large Cap Fund", "1001", 100.0, 100000, 130000, 30000, 800, 30.0, 12.5)]
    result = parse_dashboard_workbook(_write(tmp_path, {"1. Mutual Fund": rows}))
    assert result.clients[0].grand_total_reconciles is None


def test_total_rows_are_not_ingested_as_holdings(tmp_path):
    result = parse_dashboard_workbook(_simple_book(tmp_path))
    schemes = [h.scheme for h in result.clients[0].holdings]
    assert not any("total" in s.lower() for s in schemes)
    assert len(schemes) == 2


# --------------------------------------------------------------------------
# Scale shape - the counts the real files are documented to produce
# --------------------------------------------------------------------------

def test_multi_client_workbook_at_documented_scale(tmp_path):
    """Stands in for the 6-client / 311-holding file: many clients, many
    holdings, per-client attribution intact."""
    rows = [FULL_HEADERS]
    expected = {}
    for client_no in range(1, 7):
        name = f"CLIENT {client_no}"
        rows.append([f"{name} (AAAAA{client_no:04d}A)"])
        rows.append(["Equity"])
        count = 40 + client_no * 5
        expected[name] = count
        for holding_no in range(count):
            rows.append(_holding_row(
                f"Fund {client_no}-{holding_no}", f"{client_no}{holding_no:04d}",
                100.0, 100000, 130000, 30000, 800, 30.0, 12.5,
            ))
    result = parse_dashboard_workbook(_write(tmp_path, {"1. Mutual Fund": rows}))

    assert len(result.clients) == 6
    assert {c.name: len(c.holdings) for c in result.clients} == expected
    assert sum(len(c.holdings) for c in result.clients) == sum(expected.values())


# --------------------------------------------------------------------------
# The four real client exports
# --------------------------------------------------------------------------
# These are live files with real PANs and valuations, so they are
# gitignored - the tests skip when the working copies are absent rather
# than failing on a clean checkout. Every expectation below was READ OFF
# the files, not asserted in advance: where the data disagreed with what
# we expected, the file won.

REAL_DATA_DIR = Path(__file__).parent / "data"

# Files are located by GLOB and identified by the shape they parse to -
# never by filename. The real exports are named after the clients they
# belong to, so hardcoding a filename would put a client's name in this
# file just as surely as hardcoding a PAN would. The working copies are
# gitignored; these tests skip when they are absent.
#
# Each entry: (clients, holdings, workbook_total, sips). Every figure was
# READ OFF the files, not asserted in advance - where the data disagreed
# with what we expected, the file won.
REAL_FILE_SHAPES = {
    "A": (9, 311, 288_111_886.76, 95),   # the large multi-client book
    "B": (5, 57, None, 0),               # NAME-column layout, no Grand Total row
    "C": (3, 65, 44_224_401.11, 2),      # no actions at all
    "D": (2, 28, 47_454_775.03, 9),      # paired switches into one target
}


def _real_files() -> dict:
    """Maps each expectation key to the file that parses to that shape."""
    paths = sorted(REAL_DATA_DIR.glob("*.xlsx"))
    if not paths:
        pytest.skip("no real client files present in tests/data")
    by_shape = {}
    for path in paths:
        parsed = parse_dashboard_workbook(path)
        by_shape[(len(parsed.clients),
                  sum(len(c.holdings) for c in parsed.clients))] = path
    resolved = {}
    for key, (clients, holdings, _total, _sips) in REAL_FILE_SHAPES.items():
        match = by_shape.get((clients, holdings))
        if match is not None:
            resolved[key] = match
    return resolved


def _real(key):
    files = _real_files()
    if key not in files:
        clients, holdings, _t, _s = REAL_FILE_SHAPES[key]
        pytest.skip(f"no file present parsing to {clients} clients / {holdings} holdings")
    return files[key]


@pytest.mark.parametrize("key", sorted(REAL_FILE_SHAPES))
def test_real_file_shape(key):
    """Clients, holdings, and the workbook total, per real file."""
    clients, holdings, book_total, sips = REAL_FILE_SHAPES[key]
    result = parse_dashboard_workbook(_real(key))
    assert len(result.clients) == clients, [c.name for c in result.clients]
    assert sum(len(c.holdings) for c in result.clients) == holdings
    assert sum(len(c.sips) for c in result.clients) == sips
    assert result.reported_workbook_total == (
        pytest.approx(book_total) if book_total is not None else None
    )


@pytest.mark.parametrize("key", sorted(REAL_FILE_SHAPES))
def test_real_file_reconciles_to_within_one_rupee(key):
    """The reconciliation guarantee, on the actual files."""
    result = parse_dashboard_workbook(_real(key))
    if result.reported_workbook_total is not None:
        delta = abs(result.computed_workbook_total - result.reported_workbook_total)
        assert delta <= RECONCILE_TOLERANCE_RUPEES, f"workbook off by Rs {delta:,.2f}"
        assert result.workbook_total_reconciles is True
    for client in result.clients:
        if client.reported_grand_total is not None:
            assert client.grand_total_reconciles is True, (
                f"{client.name}: computed Rs {client.computed_grand_total:,.2f} vs reported "
                f"Rs {client.reported_grand_total:,.2f}"
            )


@pytest.mark.parametrize("key", sorted(REAL_FILE_SHAPES))
def test_real_file_has_no_unrecognised_actions(key):
    """Every ACTION in all four files is inside the known vocabulary. A
    new spelling must surface as a warning, never be dropped."""
    result = parse_dashboard_workbook(_real(key))
    unrecognised = [a for c in result.clients for a in c.actions if a.kind == "unrecognised"]
    assert not unrecognised, [(a.scheme, a.action_raw) for a in unrecognised]


def test_real_file_no_cagr_shift_survives_ingestion():
    """The gain-duplicated-into-CAGR defect is not present in any current
    file. This asserts the OUTCOME - no implausible CAGR is ingested -
    rather than a shift count, so it holds whether a future export
    carries the defect (repaired or rejected) or not."""
    for key in REAL_FILE_SHAPES:
        result = parse_dashboard_workbook(_real(key))
        for client in result.clients:
            for holding in client.holdings:
                if holding.cagr_pct is not None:
                    assert abs(holding.cagr_pct) <= MAX_PLAUSIBLE_CAGR_PCT, (
                        f"[{key}] {client.name} / {holding.scheme}: ingested CAGR "
                        f"{holding.cagr_pct}%"
                    )


def test_sample_file_groups_by_name_column():
    """The sample export has a NAME column and no client header rows -
    the other layout the parser has to handle.

    The expected names are READ OUT of the NAME column rather than
    written here: these are real people, and a test file is not the place
    to keep a client list. What is asserted is the property that matters -
    every distinct NAME value became exactly one client, in sorted order.
    """
    from openpyxl import load_workbook

    path = _real("B")
    sheet = load_workbook(path, read_only=True, data_only=True)["1. Mutual Fund"]
    rows = list(sheet.iter_rows(values_only=True))
    header = [str(c).strip().upper() if c else "" for c in rows[0]]
    assert "NAME" in header, "this file is expected to use NAME-column grouping"
    name_column = header.index("NAME")
    expected = sorted({
        str(r[name_column]).strip() for r in rows[1:]
        if r[name_column] and str(r[name_column]).strip()
        and not str(r[name_column]).strip().lower().endswith("total")
    })

    result = parse_dashboard_workbook(path)
    assert [c.name for c in result.clients] == expected
    assert all(c.pan is None for c in result.clients), "NAME-column layout carries no PAN"
    kinds = {a.canonical for c in result.clients for a in c.actions}
    assert {"switch", "redeem", "trim", "watchlist", "accumulate", "tax loss harvest"} == kinds


def test_jitender_file_has_zero_actions_and_still_parses():
    """A file with no actions at all must produce valid clients - the
    report builds with an empty Transaction Snapshot and no Mind Map."""
    result = parse_dashboard_workbook(_real("C"))
    assert sum(len(c.actions) for c in result.clients) == 0
    assert all(c.holdings for c in result.clients)


def test_corporate_client_is_not_mistaken_for_an_asset_class():
    """A company name in a client header row is a client, and whatever
    sits in its brackets is kept verbatim as the PAN.

    The client is located by SHAPE - a header row whose label is not in
    ASSET_CLASS_ROWS - rather than by name, so no real company name or
    PAN is written into this file. This is the case that made PAN-regex
    detection unusable: one real header carries a PAN that does not match
    the standard pattern, and validating it away would drop the client.
    """
    from openpyxl import load_workbook

    path = _real("A")
    sheet = load_workbook(path, read_only=True, data_only=True)["1. Mutual Fund"]
    header_labels = [
        str(r[0]).strip()
        for r in sheet.iter_rows(values_only=True)
        if r[0] and all(c is None for c in r[1:])
        and normalise_header(str(r[0])) not in ASSET_CLASS_ROWS
    ]
    bracketed = [lbl for lbl in header_labels if "(" in lbl and ")" in lbl]
    assert bracketed, "expected client header rows of the form 'NAME (PAN)'"

    result = parse_dashboard_workbook(path)
    parsed_names = {c.name for c in result.clients}
    for label in bracketed:
        name, pan = split_client_label(label)
        assert name in parsed_names, f"client header {name!r} did not become a client"
        client = result.client(name)
        assert client.pan == pan, "the bracketed value must be kept verbatim"
        assert client.holdings, f"{name!r} was created with no holdings"


# --------------------------------------------------------------------------
# Switch pairing (assembler)
# --------------------------------------------------------------------------

def test_two_switches_into_the_same_target_each_get_their_own_switch_in():
    """One client in the two-client book switches two folios of the same fund
    into the same target. Pairing by target-scheme name collapsed both
    onto one Switch In row - one was left with no amount, the other
    overwritten, and half the money disappeared from the report.

    The client is found by that SHAPE (two switch actions sharing a
    suggested scheme) rather than by name.
    """
    from collections import Counter
    from datetime import date
    from pipeline.report_assembler import assemble_report_context
    from pipeline.docx_builder import RMInfo, _validate_report_context

    result = parse_dashboard_workbook(_real("D"))
    target = None
    for candidate in result.clients:
        switches = [a for a in candidate.actions if a.canonical == "switch"]
        repeated = [t for t, n in Counter(a.suggested_scheme for a in switches).items() if n > 1]
        if repeated:
            target = candidate
            break
    assert target is not None, "expected a client switching twice into one target scheme"

    rm = RMInfo(name="RM", email="rm@wealthcareindia.com", phone="+91")
    assembled = assemble_report_context(client=target, as_of=date.today(), rm=rm)
    rows = assembled.ctx.transaction_snapshot

    outs = [t for t in rows if t.action == "Switch Out"]
    ins = [t for t in rows if t.action == "Switch In"]
    assert len(outs) == 2 and len(ins) == 2, [(t.action, t.scheme) for t in rows]
    assert all(t.amount is not None for t in ins), \
        "a Switch In was left with no amount - its Switch Out funded a different row"

    out_total = sum(t.amount for t in outs)
    in_total = sum(t.amount for t in ins)
    deductions = sum(t.switch_deduction or 0.0 for t in outs)
    assert round(out_total - deductions - in_total, 2) == 0, (
        f"Rs {out_total - deductions - in_total:,.2f} unaccounted between switch out and in"
    )
    # The build-time guard must agree.
    _validate_report_context(assembled.ctx)


@pytest.mark.parametrize("key", sorted(REAL_FILE_SHAPES))
def test_every_real_client_passes_the_build_time_invariants(key):
    """Assemble every client in every real file and run the report
    context through the same validation build_report() applies. This is
    what caught the switch-pairing bug."""
    from datetime import date
    from pipeline.report_assembler import assemble_report_context
    from pipeline.docx_builder import RMInfo, _validate_report_context

    result = parse_dashboard_workbook(_real(key))
    rm = RMInfo(name="RM", email="rm@wealthcareindia.com", phone="+91")
    for client in result.clients:
        assembled = assemble_report_context(client=client, as_of=date.today(), rm=rm)
        _validate_report_context(assembled.ctx)
